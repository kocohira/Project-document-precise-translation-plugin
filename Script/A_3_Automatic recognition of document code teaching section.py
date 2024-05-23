#! /usr/bin/env python3
#! /usr/bin/env python
import goto
from dominate.tags import label
from goto import with_goto

@with_goto
def Total() :

    label.begin

    try :

        from selenium import webdriver
        import json
        import re       ###正则表达式用
        from openpyxl import load_workbook          ###Excel写入用
        from datetime import datetime
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import nltk                                 ###自然语言处理用
        import time,win32api,win32con
        import os
        import sys



        def restart_program() :

            print('重启程序中...')

            python = sys.executable

            os.execl(python,python,*sys.argv)




        #实例化谷歌设置选项
        option = webdriver.ChromeOptions()
        #添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
        #option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
        option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
        #option.add_argument("--headless")

        experimentalFlags = [
            #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
            'same-site-by-default-cookies@2',
            #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
            'cookies-without-same-site-must-be-secure@2',]
        chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
        option.add_experimental_option('localState', chromeLocalStatePrefs)



        #初始化driver
        driver = webdriver.Chrome(options=option)








        Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

        Sheet = Excel['大阪']  ###打开‘大阪’工作表

        Count_Row = int((Sheet.cell(row=1, column=31)).value)                             ###每次出错就从这里修改

        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)                ###搜索框中要输入的店铺

        Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)                  ###搜索框中要输入的区域




        ########Excel计数插件

        Total_Row = 1

        Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        while Total_Quantity != None :

            Total_Row = Total_Row + 1

            Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        Max_Row = Total_Row - 1             ###包含了表格第一行

        print(Max_Row)

        ########Excel计数插件








        url = 'https://tabelog.com/osaka/A2701/A270405/27011913/'



        driver.get(url)

        cookies2_txt = """[
        {
            "domain": ".dianping.com",
            "expirationDate": 1668961275,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_ga",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.2.723816111.1603966407",
            "id": 1
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1605889335,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_gat_gtag_UA_111805100_1",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 2
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1605975675,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_gid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.2.369363227.1605818515",
            "id": 3
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1635502415,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_hc.v",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "25c06052-ec24-b321-75e2-df2018908df6.1603966415",
            "id": 4
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606235459,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lx_utm",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "utm_source%3Dgoogle%26utm_medium%3Dorganic",
            "id": 5
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1698574407,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
            "id": 6
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1698574407,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk_cuid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
            "id": 7
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1605890141,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk_s",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "175e6599fa8-5fc-f4c-324%7C%7C3",
            "id": 8
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094172,
            "hostOnly": false,
            "httpOnly": false,
            "name": "aburl",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 9
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606486817.735597,
            "hostOnly": false,
            "httpOnly": false,
            "name": "bsid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "KIm9nv3SR27k0fzg4zPvth8BIQcTEZwQUs_66vzjsOk4aFXKveYCZ8G0Eqb2E-QlrLHI_s_guYaocm60ZQ3UTg",
            "id": 10
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1635761147.17086,
            "hostOnly": false,
            "httpOnly": false,
            "name": "ctu",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "9f07eccced25cf82543a98c2d6e0235399e464f43265dd96bf67e9790d7d49ed",
            "id": 11
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094249,
            "hostOnly": false,
            "httpOnly": false,
            "name": "cy",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "2374",
            "id": 12
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094289,
            "hostOnly": false,
            "httpOnly": false,
            "name": "cye",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "osaka",
            "id": 13
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606486817.735668,
            "hostOnly": false,
            "httpOnly": false,
            "name": "edper",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "KIm9nv3SR27k0fzg4zPvth8BIQcTEZwQUs_66vzjsOk4aFXKveYCZ8G0Eqb2E-QlrLHI_s_guYaocm60ZQ3UTg",
            "id": 14
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659,
            "hostOnly": false,
            "httpOnly": false,
            "name": "Hm_lvt_602b80cf8079ae6591966cc70a3940e7",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1604647683,1604835193,1605286882,1605628795",
            "id": 15
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1667994073.957559,
            "hostOnly": false,
            "httpOnly": false,
            "name": "s_ViewType",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "10",
            "id": 16
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1636371292.046143,
            "hostOnly": false,
            "httpOnly": false,
            "name": "ua",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "dpuser_83084044340",
            "id": 17
        },
        {
            "domain": ".eo.dianping.com",
            "expirationDate": 1606494075,
            "hostOnly": false,
            "httpOnly": false,
            "name": "os-agent",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "",
            "id": 18
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "JSESSIONID",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "E72C3035785E5032D843FDF05CA009A7",
            "id": 19
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "logan_custom_report",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "",
            "id": 20
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "logan_session_token",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "fx6olnip3eu2qra64wfa",
            "id": 21
        }
        ]"""          ###cookies内容

        cookies2 = json.loads(cookies2_txt)###cookies json化

        driver.delete_all_cookies()      ###清除原有的cookies

        for cook in cookies2 :            ###cookies去sameSite开始
            try :
                cook.pop('sameSite')
            except :
                pass
            driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址
        driver.get(url)                  ###重新加载网址








        while Count_Row < Max_Row + 1 :

            WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='sk']")))          ###等待出现搜索店铺的搜索框加载完毕

            driver.find_element_by_xpath("//*[@id='sk']").send_keys(Restanrant_Search)                          ###找到输入店铺的搜索框，输入店铺

            WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='sa']")))          ###等待出现搜索店铺的搜索框加载完毕

            driver.find_element_by_xpath("//*[@id='sa']").send_keys(Location_Search)                            ###找到输入区域的搜索框，输入区域

            WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='js-global-search-btn']")))          ###等待出现搜索店铺的搜索框加载完毕

            driver.find_element_by_xpath("//*[@id='js-global-search-btn']").click()                             ###找到搜索按钮并点击，窗口跳到下一个画面，但控制的页面没变

            time.sleep(2.5)




            Location_Judgement_1 = driver.find_elements_by_xpath("//*[contains(text(),'ご指定の条件に該当するお店は見つかりませんでした。')]")             ###第一次判断有无店铺

            if len(Location_Judgement_1) == 0 :

                Location_Judgement_2 = driver.find_elements_by_xpath("//*[contains(text(),'検索結果が0件のため、')]")                                 ###第二次判断有无店铺

                if len(Location_Judgement_2) == 0 :

                    WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@class='list-controll clearfix']/following-sibling::div[1]/div[1]")))            ###等待第一个搜索结果加载完毕

                    time.sleep(2.5)

                    driver.find_element_by_xpath("//*[@class='list-controll clearfix']/following-sibling::div[1]/div[1]").click()         ###

                    time.sleep(1.5)

                    windows = driver.window_handles                                                                     ####搜索当前所有打开的页面的句柄（识别用）

                    driver.close()                                                                                      ###关闭上一个页面

                    driver.switch_to.window(windows[-1])                                                                ###切换到要控制的新开的网页的句柄。





                    ###在这里判断店铺有无电话，地址，营业时间

                    TBRG_Telephone_Pre_Judgement = driver.find_elements_by_xpath("//*[@class='rstinfo-table__tel-num-wrap']")

                    if len(TBRG_Telephone_Pre_Judgement) != 0 :

                        TBRG_Telephone = driver.find_element_by_xpath("//*[@class='rstinfo-table__tel-num-wrap']").text

                        TBRG_Telephone = re.sub("\D", "", TBRG_Telephone)

                        print(TBRG_Telephone)

                        TBRG_Telephone_Judgement = TBRG_Telephone.isnumeric()

                        if TBRG_Telephone_Judgement == True :

                            TBRG_Location_Pre_Judgement = driver.find_elements_by_xpath("//*[@class='rstinfo-table__address']")

                            if len(TBRG_Location_Pre_Judgement) != 0 :

                                TBRG_Location = driver.find_element_by_xpath("//*[@class='rstinfo-table__address']").text

                                TBRG_Times = driver.find_elements_by_xpath("//*[@class='rstinfo-table__subject' and text()='営業時間']/following-sibling::p[1]")

                                if len(TBRG_Times) != 0 :

                                    TBRG_Weeks = driver.find_elements_by_xpath("//*[@class='rstinfo-table__subject' and text()='定休日']/following-sibling::p[1]")

                                    if len(TBRG_Weeks) != 0 :

                                        TBRG_Time = driver.find_element_by_xpath("//*[@class='rstinfo-table__subject' and text()='営業時間']/following-sibling::p[1]").text             ####疫情结束前都会有这个Xpath

                                        TBRG_Week = driver.find_element_by_xpath("//*[@class='rstinfo-table__subject' and text()='定休日']/following-sibling::p[1]").text              ####疫情结束前都会有这个Xpath

                                        TBRG_Location_Judgement = TBRG_Location

                                        TBRG_Time_Judgement = TBRG_Time

                                        TBRG_Week_Judgement = TBRG_Week

                                        print(TBRG_Location)

                                        print(TBRG_Time)

                                        print(TBRG_Week)



                                        if len(TBRG_Location_Judgement) > 5 and len(TBRG_Time_Judgement) > 5:

                                            Display_Name = driver.find_element_by_xpath("//*[@class='display-name']")

                                            Display_Name = Display_Name.text

                                            Element_Quantity_1 = driver.find_elements_by_xpath("//*[@class='homepage']")

                                            Element_Exist_Judgement_1 = len(Element_Quantity_1)

                                            if Element_Exist_Judgement_1 == 0 :

                                                Element_Quantity_2 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-facebook']")

                                                Element_Exist_Judgement_2 = len(Element_Quantity_2)

                                                if Element_Exist_Judgement_2 == 0 :

                                                    Element_Quantity_3 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-twitter']")

                                                    Element_Exist_Judgement_3 = len(Element_Quantity_3)

                                                    if Element_Exist_Judgement_3 == 0 :

                                                        Element_Quantity_4 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-instagram']")

                                                        Element_Exist_Judgement_4 = len(Element_Quantity_4)

                                                        if Element_Exist_Judgement_4 == 0 :

                                                            print('No TBRG_Website')

                                                        else :

                                                            TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-instagram']")

                                                            Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                            print(TBRG_Website.text)

                                                    else :

                                                        TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-twitter']")

                                                        Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                        print(TBRG_Website.text)

                                                else :

                                                    TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-facebook']")

                                                    Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                    print(TBRG_Website.text)

                                            else :

                                                TBRG_Website = driver.find_element_by_xpath("//*[@class='homepage']")

                                                Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                print(TBRG_Website.text)








                                            Booksite_Quantity_1 = driver.find_elements_by_xpath("//*[text()='予約する']")

                                            Booksite_Exist_Judgement_1 = len(Booksite_Quantity_1)

                                            if Booksite_Exist_Judgement_1 == 0 :

                                                Booksite_Quantity_2 = driver.find_elements_by_xpath("//*[text()='予約申し込み']")

                                                Booksite_Exist_Judgement_2 = len(Booksite_Quantity_2)

                                                if Booksite_Exist_Judgement_2 == 0 :

                                                    print('No TBLG_Booksite')

                                                else :

                                                    Origin_Windle_Handle = driver.current_window_handle

                                                    #time.sleep(2.5)

                                                    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'予約申し込み')]")))  ###等待出现搜索店铺的搜索框加载完毕

                                                    #_item = driver.find_element_by_class_name('')



                                                    _item = driver.find_element_by_link_text('予約申し込み')

                                                    driver.execute_script("arguments[0].click();", _item)

                                                    All_Window_Handles = driver.window_handles

                                                    for handle in All_Window_Handles :

                                                        if handle != Origin_Windle_Handle :

                                                            driver.switch_to.window(handle)

                                                    time.sleep(3)

                                                    #driver.find_element_by_xpath("//*[@class='shopName']").click()

                                                    win32api.keybd_event(27, 0, 0, 0)

                                                    win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

                                                    #New_Windle_Handle = driver.current_window_handle

                                                    #driver.switch_to.window(New_Windle_Handle)

                                                    TBLG_Booksite = driver.current_url

                                                    driver.close()

                                                    driver.switch_to.window(Origin_Windle_Handle)

                                                    Sheet.cell(row=Count_Row, column=13).value = TBLG_Booksite

                                                    print(TBLG_Booksite)

                                            else :

                                                win32api.keybd_event(27, 0, 0, 0)

                                                win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

                                                TBLG_Booksite = driver.current_url

                                                Sheet.cell(row=Count_Row, column=13).value = TBLG_Booksite

                                                print(TBLG_Booksite)




                                            if len(TBRG_Week_Judgement) != 0:

                                                Sheet.cell(row=Count_Row, column=23).value = TBRG_Week

                                                Sheet.cell(row=Count_Row, column=11).value = TBRG_Location

                                                Sheet.cell(row=Count_Row, column=12).value = TBRG_Telephone

                                                Sheet.cell(row=Count_Row, column=5).value = re.sub("[]", " ", TBRG_Time)

                                                Sheet.cell(row=Count_Row, column=2).value = Display_Name





                                            else:

                                                Sheet.cell(row=Count_Row, column=23).value = '無休'

                                                Sheet.cell(row=Count_Row, column=11).value = TBRG_Location

                                                Sheet.cell(row=Count_Row, column=12).value = TBRG_Telephone

                                                Sheet.cell(row=Count_Row, column=5).value = re.sub("[]", " ", TBRG_Time)

                                                Sheet.cell(row=Count_Row, column=2).value = Display_Name




                                            print(Count_Row)

                                            Count_Row = Count_Row + 1

                                            Sheet.cell(row=1, column=31).value = Count_Row

                                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                            Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                            driver.get(url)

                                            time.sleep(2.5)

                                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                            #driver.close()  ###关闭上一个页面

                                            driver.switch_to.window(windows[-1])

                                            time.sleep(2.5)


                                        else :

                                            if len(TBRG_Location_Judgement) <= 5 :

                                                Sheet.cell(row=Count_Row, column=11).value = '####'

                                                print(Count_Row)

                                                Count_Row = Count_Row + 1

                                                Sheet.cell(row=1, column=31).value = Count_Row

                                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                                driver.get(url)

                                                time.sleep(2.5)

                                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                                #driver.close()  ###关闭上一个页面

                                                driver.switch_to.window(windows[-1])

                                                time.sleep(2.5)


                                            if len(TBRG_Time_Judgement) <= 5 :

                                                Sheet.cell(row=Count_Row, column=5).value = '####'

                                                print(Count_Row)

                                                Count_Row = Count_Row + 1

                                                Sheet.cell(row=1, column=31).value = Count_Row

                                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                                driver.get(url)

                                                time.sleep(2.5)

                                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                                #driver.close()  ###关闭上一个页面

                                                driver.switch_to.window(windows[-1])

                                                time.sleep(2.5)




                                    else :

                                        TBRG_Time = driver.find_element_by_xpath("//*[@class='rstinfo-table__subject' and text()='営業時間']/following-sibling::p[1]").text  ####疫情结束前都会有这个Xpath

                                        #TBRG_Week = driver.find_element_by_xpath("//*[@class='rstinfo-table__subject' and text()='定休日']/following-sibling::p[1]").text  ####疫情结束前都会有这个Xpath

                                        TBRG_Location_Judgement = TBRG_Location

                                        TBRG_Time_Judgement = TBRG_Time

                                        #TBRG_Week_Judgement = TBRG_Week

                                        print(TBRG_Location)

                                        print(TBRG_Time)

                                        #print(TBRG_Week)

                                        if len(TBRG_Location_Judgement) > 5 and len(TBRG_Time_Judgement) > 5:

                                            Display_Name = driver.find_element_by_xpath("//*[@class='display-name']")

                                            Display_Name = Display_Name.text

                                            Element_Quantity_1 = driver.find_elements_by_xpath("//*[@class='homepage']")

                                            Element_Exist_Judgement_1 = len(Element_Quantity_1)

                                            if Element_Exist_Judgement_1 == 0:

                                                Element_Quantity_2 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-facebook']")

                                                Element_Exist_Judgement_2 = len(Element_Quantity_2)

                                                if Element_Exist_Judgement_2 == 0:

                                                    Element_Quantity_3 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-twitter']")

                                                    Element_Exist_Judgement_3 = len(Element_Quantity_3)

                                                    if Element_Exist_Judgement_3 == 0:

                                                        Element_Quantity_4 = driver.find_elements_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-instagram']")

                                                        Element_Exist_Judgement_4 = len(Element_Quantity_4)

                                                        if Element_Exist_Judgement_4 == 0:

                                                            print('No TBRG_Website')

                                                        else:

                                                            TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-instagram']")

                                                            Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                            print(TBRG_Website.text)

                                                    else:

                                                        TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-twitter']")

                                                        Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                        print(TBRG_Website.text)

                                                else:

                                                    TBRG_Website = driver.find_element_by_xpath("//*[@class='rstinfo-sns-link rstinfo-sns-facebook']")

                                                    Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                    print(TBRG_Website.text)

                                            else:

                                                TBRG_Website = driver.find_element_by_xpath("//*[@class='homepage']")

                                                Sheet.cell(row=Count_Row, column=14).value = TBRG_Website.text

                                                print(TBRG_Website.text)




                                            Booksite_Quantity_1 = driver.find_elements_by_xpath("//*[text()='予約する']")

                                            Booksite_Exist_Judgement_1 = len(Booksite_Quantity_1)

                                            if Booksite_Exist_Judgement_1 == 0:

                                                Booksite_Quantity_2 = driver.find_elements_by_xpath("//*[text()='予約申し込み']")

                                                Booksite_Exist_Judgement_2 = len(Booksite_Quantity_2)

                                                if Booksite_Exist_Judgement_2 == 0:

                                                    print('No TBLG_Booksite')

                                                else:

                                                    Origin_Windle_Handle = driver.current_window_handle

                                                    # time.sleep(2.5)

                                                    WebDriverWait(driver, 50).until(EC.presence_of_element_located((By.XPATH, "//*[contains(text(),'予約申し込み')]")))  ###等待出现搜索店铺的搜索框加载完毕

                                                    # _item = driver.find_element_by_class_name('')

                                                    _item = driver.find_element_by_link_text('予約申し込み')

                                                    driver.execute_script("arguments[0].click();", _item)

                                                    All_Window_Handles = driver.window_handles

                                                    for handle in All_Window_Handles:

                                                        if handle != Origin_Windle_Handle:

                                                            driver.switch_to.window(handle)

                                                    time.sleep(3)

                                                    # driver.find_element_by_xpath("//*[@class='shopName']").click()

                                                    win32api.keybd_event(27, 0, 0, 0)

                                                    win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

                                                    # New_Windle_Handle = driver.current_window_handle

                                                    # driver.switch_to.window(New_Windle_Handle)

                                                    TBLG_Booksite = driver.current_url

                                                    driver.close()

                                                    driver.switch_to.window(Origin_Windle_Handle)

                                                    Sheet.cell(row=Count_Row, column=13).value = TBLG_Booksite

                                                    print(TBLG_Booksite)

                                            else:

                                                win32api.keybd_event(27, 0, 0, 0)

                                                win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

                                                TBLG_Booksite = driver.current_url

                                                Sheet.cell(row=Count_Row, column=13).value = TBLG_Booksite

                                                print(TBLG_Booksite)





                                            Sheet.cell(row=Count_Row, column=23).value = '無休'

                                            Sheet.cell(row=Count_Row, column=11).value = TBRG_Location

                                            Sheet.cell(row=Count_Row, column=12).value = TBRG_Telephone

                                            Sheet.cell(row=Count_Row, column=5).value = re.sub("[]", " ", TBRG_Time)

                                            Sheet.cell(row=Count_Row, column=2).value = Display_Name

                                            print(Count_Row)

                                            Count_Row = Count_Row + 1

                                            Sheet.cell(row=1, column=31).value = Count_Row

                                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                            Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                            driver.get(url)

                                            time.sleep(2.5)

                                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                            # driver.close()  ###关闭上一个页面

                                            driver.switch_to.window(windows[-1])

                                            time.sleep(2.5)


                                        else:

                                            if len(TBRG_Location_Judgement) <= 5:

                                                Sheet.cell(row=Count_Row, column=11).value = '####'

                                                print(Count_Row)

                                                Count_Row = Count_Row + 1

                                                Sheet.cell(row=1, column=31).value = Count_Row

                                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                                driver.get(url)

                                                time.sleep(2.5)

                                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                                # driver.close()  ###关闭上一个页面

                                                driver.switch_to.window(windows[-1])

                                                time.sleep(2.5)

                                            if len(TBRG_Time_Judgement) <= 5:

                                                Sheet.cell(row=Count_Row, column=5).value = '####'

                                                print(Count_Row)

                                                Count_Row = Count_Row + 1

                                                Sheet.cell(row=1, column=31).value = Count_Row

                                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                                driver.get(url)

                                                time.sleep(2.5)

                                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                                # driver.close()  ###关闭上一个页面

                                                driver.switch_to.window(windows[-1])

                                                time.sleep(2.5)

                                else :

                                    Sheet.cell(row=Count_Row, column=5).value = '####'

                                    print(Count_Row)

                                    Count_Row = Count_Row + 1

                                    Sheet.cell(row=1, column=31).value = Count_Row

                                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                    Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                    driver.get(url)

                                    time.sleep(2.5)

                                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                    # driver.close()  ###关闭上一个页面

                                    driver.switch_to.window(windows[-1])

                                    time.sleep(2.5)

                            else :

                                Sheet.cell(row=Count_Row, column=11).value = '####'

                                print(Count_Row)

                                Count_Row = Count_Row + 1

                                Sheet.cell(row=1, column=31).value = Count_Row

                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                                driver.get(url)

                                time.sleep(2.5)

                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                # driver.close()  ###关闭上一个页面

                                driver.switch_to.window(windows[-1])

                                time.sleep(2.5)

                        else :

                            Sheet.cell(row=Count_Row, column=12).value = '####'

                            print(Count_Row)

                            Count_Row = Count_Row + 1

                            Sheet.cell(row=1, column=31).value = Count_Row

                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                            Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                            driver.get(url)

                            time.sleep(2.5)

                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                            #driver.close()  ###关闭上一个页面

                            driver.switch_to.window(windows[-1])

                            time.sleep(2.5)

                    else :

                        Sheet.cell(row=Count_Row, column=12).value = '####'

                        print(Count_Row)

                        Count_Row = Count_Row + 1

                        Sheet.cell(row=1, column=31).value = Count_Row

                        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                        Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                        driver.get(url)

                        time.sleep(2.5)

                        windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                        # driver.close()  ###关闭上一个页面

                        driver.switch_to.window(windows[-1])

                        time.sleep(2.5)

                else :

                    Sheet.cell(row=Count_Row, column=11).value = '####'

                    print(Count_Row)

                    Count_Row = Count_Row + 1

                    Sheet.cell(row=1, column=31).value = Count_Row

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                    Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                    driver.get(url)

                    time.sleep(2.5)

                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                    #driver.close()  ###关闭上一个页面

                    driver.switch_to.window(windows[-1])

                    time.sleep(2.5)

            else :

                Sheet.cell(row=Count_Row, column=11).value = '####'

                print(Count_Row)

                Count_Row = Count_Row + 1

                Sheet.cell(row=1, column=31).value = Count_Row

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)  ###搜索框中要输入的区域

                driver.get(url)

                time.sleep(2.5)

                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                #driver.close()  ###关闭上一个页面

                driver.switch_to.window(windows[-1])

                time.sleep(2.5)




        ########时间除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=5).value

        while Count_Row < Max_Row + 1 :

            if None_Count == '####' :

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=5).value




                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:

                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)




            if None_Count != '####' :

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=5).value

        time.sleep(0.3)

        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")

        ########时间除空插件




        ########地址除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=11).value

        while Count_Row < Max_Row + 1 :

            if None_Count == '####' :

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=11).value




                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:

                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)




            if None_Count != '####' :

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=11).value

        time.sleep(0.3)

        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")

        ########地址除空插件




        ########电话除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=12).value

        while Count_Row < Max_Row + 1 :

            if None_Count == '####' :

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=12).value




                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:

                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)




            if None_Count != '####' :

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=12).value

        time.sleep(0.3)

        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")

        ########电话除空插件








    except :

        # ! /usr/bin/env python3
        # ! /usr/bin/env python

        #import time,win32api,win32con
        #import os
        #import sys
        #import 3_A_食べログ餐厅信息搜寻系统

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        driver.quit()

        #time.sleep(8)

        #goto.begin

        #def restart_program() :

            #print('重启程序中...')

            #python = sys.executable

            #os.execl(python,python,*sys.argv)

        #os.popen('D:\应用方面\python project2\A_3_食べログ餐厅信息搜寻系统.py')
        #restart_program()




Total()
###NT








