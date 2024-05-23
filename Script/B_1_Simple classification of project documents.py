from selenium import webdriver
import time
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用




Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表

######BUG修改区

#Count = 1

#Input_Row = 2

Count = 1

Input_Row = 2

######BUG修改区




######计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1

print(Max_Row)




######计数插件




Count_Row = 2

None_Count = Sheet.cell(row=Count_Row, column=2).value

while Count_Row < Max_Row + 2 :

    Judgement_1 = None_Count.find('冰')

    Judgement_2 = None_Count.find('饼')

    Judgement_3 = None_Count.find('麦当劳')

    Judgement_4 = None_Count.find('星巴克')

    Judgement_5 = None_Count.find('肯德基')

    Judgement_6 = None_Count.find('糕')

    Judgement_7 = None_Count.find('咖啡')

    Judgement_8 = None_Count.find('Pizza')

    Judgement_9 = None_Count.find('披萨')

    Judgement_10 = None_Count.find('Dessert')

    Judgement_11 = None_Count.find('点心')

    Judgement_12 = None_Count.find('餅')

    Judgement_13 = None_Count.find('披薩')

    Judgement_14 = None_Count.find('點心')

    Judgement_15 = None_Count.find('cafe')

    Judgement_16 = None_Count.find('Cafe')

    Judgement_17 = None_Count.find('CAFE')

    Judgement_18 = None_Count.find('Caffé')

    Judgement_19 = None_Count.find('Café')

    Judgement_20 = None_Count.find('Coffee')

    Judgement_21 = None_Count.find('Cake')

    Judgement_22 = None_Count.find('cake')

    Judgement_23 = None_Count.find('Haagen-Dazs')

    Judgement_24 = None_Count.find('哈根达斯')

    Judgement_25 = None_Count.find('哈根達斯')


    if Judgement_1 == -1 and Judgement_2 == -1 and Judgement_3 == -1 and Judgement_4 == -1 and Judgement_5 == -1 and Judgement_6 == -1 and Judgement_7 == -1 and Judgement_8 == -1 and Judgement_9 == -1 and Judgement_10 == -1 and Judgement_11 == -1 and Judgement_12 == -1 and Judgement_13 == -1 and Judgement_14 == -1 and Judgement_15 == -1 and Judgement_16 == -1 and Judgement_17 == -1 and Judgement_18 == -1 and Judgement_19 == -1 and Judgement_20 == -1 and Judgement_21 == -1 and Judgement_22 == -1 and Judgement_23 == -1 and Judgement_24 == -1 and Judgement_25 == -1:

        print(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=2).value




    else :

        print(Count_Row)

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        #Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=2).value




time.sleep(0.3)

print("筛选结束。")
print("筛选结束。")
print("筛选结束。")
print("筛选结束。")
print("筛选结束。")
print("筛选结束。")
print("筛选结束。")
print("筛选结束。")






