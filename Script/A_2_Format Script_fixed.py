from selenium import webdriver
import time
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用



Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表

Excel_Dictionary = load_workbook(r'C:\Users\kocohira\Desktop\Special\辖区&日期_字典.xlsx') ###打开字典Excel

Sheet_Dictionary = Excel_Dictionary['字典']                                              ###打开‘字典’工作表

Row1 = 2                                                                                ###设定读取字典的起始行数

Column1 = 1                                                                             ###设定读取字典的起始列数

Row2 = 2                                                                                ###设定读取MT餐厅地址的起始行数

Column2 = 8                                                                             ###设定读取MT餐厅地址的起始列数

Location_Dictionary = str((Sheet_Dictionary.cell(row=Row1, column=Column1)).value)      ###读取字典元素，并将其字符串格式化

Location_MT = str((Sheet.cell(row=Row2, column=Column2)).value)                         ###读取MT餐厅地址，并将其字符串格式化




########Excel计数插件A

Total_Row_Sheet = 1

Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

while Total_Quantity != None :

    Total_Row_Sheet = Total_Row_Sheet + 1

    Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

Max_Row_Sheet = Total_Row_Sheet - 1

print(Max_Row_Sheet)

########Excel计数插件A




########Excel计数插件B(字典)

Total_Row_Sheet_Dictionary = 1

Total_Quantity = Sheet_Dictionary.cell(row=Total_Row_Sheet_Dictionary, column=1).value

while Total_Quantity != None :

    Total_Row_Sheet_Dictionary = Total_Row_Sheet_Dictionary + 1

    Total_Quantity = Sheet_Dictionary.cell(row=Total_Row_Sheet_Dictionary, column=1).value

Max_Row_Sheet_Dictionary = Total_Row_Sheet_Dictionary - 1

print(Max_Row_Sheet_Dictionary)

########Excel计数插件B(字典)




while Row2 < Max_Row_Sheet + 1 :                                                                        ###这里应该MT最大餐厅地址数在Excel里的行数为上限

    while Location_MT.find(Location_Dictionary) == -1 and Row1 < Max_Row_Sheet_Dictionary :                   ###比较MT餐厅地址里有无当前字典元素，以及是否比较完毕（共76个）

        Row1 = Row1 + 1                                                                  ###无当前字典元素，行数+1向下走

        Location_Dictionary = (Sheet_Dictionary.cell(row=Row1, column=Column1)).value   ###刷新字典元素

    if Row1 >= Max_Row_Sheet_Dictionary :                                                                      ###当比较完毕时（77》76）进入if语句

        print('####')

        Sheet.cell(row=Row2, column=9).value = '####'                                   ###在对应MT餐厅地址的输出结果一列填入‘空’

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Row1 = 2                                                                        ###复原字典行数

        Row2 = Row2 + 1                                                                 ###MT餐厅地址行数+1，即进入下一家餐厅的比对

        Location_Dictionary = (Sheet_Dictionary.cell(row=Row1, column=Column1)).value   ###刷新字典元素

        Location_MT = (Sheet.cell(row=Row2, column=Column2)).value                      ###刷新MT餐厅地址

    else :                                                                              ###当比较完毕时（其中一个字典元素与MT餐厅地址比对成功）进入else语句

        print(Sheet_Dictionary.cell(row=Row1, column=3).value)

        Sheet.cell(row=Row2, column=9).value = Sheet_Dictionary.cell(row=Row1, column=3).value          ###在对应MT餐厅地址的输出结果一列填入字典元素对应的输出结果

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Row1 = 2                                                                                        ###复原字典行数

        Row2 = Row2 + 1                                                                                 ###MT餐厅地址行数+1，即进入下一家餐厅的比对

        Location_Dictionary = (Sheet_Dictionary.cell(row=Row1, column=Column1)).value                   ###刷新字典元素

        Location_MT = (Sheet.cell(row=Row2, column=Column2)).value                                      ###刷新MT餐厅地址

print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')




########除空插件

Count_Row = 1

None_Count = Sheet.cell(row=Count_Row, column=9).value

while Count_Row < Max_Row_Sheet + 1 :

    if None_Count == '####' :

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row

        None_Count = Sheet.cell(row=Count_Row, column=9).value




        Total_Row_Sheet = 1

        Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        while Total_Quantity != None:

            Total_Row_Sheet = Total_Row_Sheet + 1

            Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        Max_Row_Sheet = Total_Row_Sheet - 1

        print(Max_Row_Sheet)




    if None_Count != '####' :

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=9).value

time.sleep(0.3)

print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")


########除空插件###NT





