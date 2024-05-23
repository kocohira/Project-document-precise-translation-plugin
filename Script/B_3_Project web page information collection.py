from selenium import webdriver
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import nltk                                 ###自然语言处理用
from nltk.tokenize import word_tokenize
import time,win32api,win32con
import os
import sys
from selenium.webdriver.common.keys import Keys






#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
#option.add_argument("--headless")

experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)



#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

Sheet = Excel['大阪']  ###打开‘大阪’工作表

Count_Row = int((Sheet.cell(row=1, column=31)).value)                             ###每次出错就从这里修改

Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)                ###搜索框中要输入的店铺






########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件




url = 'https://www.google.com'

driver.get(url)

cookies2_txt = '''[
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712625,
    "hostOnly": false,
    "httpOnly": false,
    "name": "__Secure-3PAPISID",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "U7Jvl7_7Sluz0hBG/A6Am1zdaoueAHRNbq",
    "id": 1
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712126,
    "hostOnly": false,
    "httpOnly": true,
    "name": "__Secure-3PSID",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "4QdypGEtRQ9ATOo97TQkns3_Qm0FTjIQPwMwblSRZb1rAElh9BI279KuxIfkJcPPWL8TsA.",
    "id": 2
},
{
    "domain": ".google.com",
    "expirationDate": 1639807959.819438,
    "hostOnly": false,
    "httpOnly": true,
    "name": "__Secure-3PSIDCC",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "AJi4QfETeFob6COusZRjGXnLoZh7sBCN-7AfbPegTkleU33nzx1rYe8IUPYWg7DeYEohgkKHsJA",
    "id": 3
},
{
    "domain": ".google.com",
    "expirationDate": 1610863959.818455,
    "hostOnly": false,
    "httpOnly": false,
    "name": "1P_JAR",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "2020-12-18-06",
    "id": 4
},
{
    "domain": ".google.com",
    "expirationDate": 1636489908.791502,
    "hostOnly": false,
    "httpOnly": true,
    "name": "ANID",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "AHWqTUlDTcY5iQ1szdjO7bwe8Acnfc_2tR07fYhG_-nb2FN5kf2yXxhZzDBASnmT",
    "id": 5
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712472,
    "hostOnly": false,
    "httpOnly": false,
    "name": "APISID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "z8bR09NnYVj2SAt0/ADdd56vIgnAEzaR5l",
    "id": 6
},
{
    "domain": ".google.com",
    "expirationDate": 2146723199.614988,
    "hostOnly": false,
    "httpOnly": false,
    "name": "CONSENT",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "YES+GB.zh-CN+V14+BX",
    "id": 7
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712342,
    "hostOnly": false,
    "httpOnly": true,
    "name": "HSID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "AbyPQoXKMcMQkQrIb",
    "id": 8
},
{
    "domain": ".google.com",
    "expirationDate": 1624083151.551027,
    "hostOnly": false,
    "httpOnly": true,
    "name": "NID",
    "path": "/",
    "sameSite": "no_restriction",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "205=ii0rVR3DswnDTEDRDfNGJ6CmDBn2eIWK91lHtzwUy7l8cTE9OdLexKVllGSdzqRWvjSQqincCvvrHp3Ws5CfDkbaKJbElJW7sVlwTNQdTbarXIiDcPZyPFfsS9th7rp91jKt45djUzfogvI237-IhlfheFxdFrsfsmAvS04CPnRl6Yuj9pOxF1mMOH9TdwdkVC9W2bS3MLt3_w3PbSNpMyvkeC9PK0ZkYc-klcj5T3vJS_aG4te7SvRzDxZGCi5jlG4yD0Pafqc8nWr9rigeQYbZMyaZEfPf2P-1fpApsxTQrrJw-5xJX5A5WrWxHqMh2_5Ue8wQBqvG7dDHbb_6ghcAPgU1ioxuN8VgNP8",
    "id": 9
},
{
    "domain": ".google.com",
    "expirationDate": 1610709692,
    "hostOnly": false,
    "httpOnly": false,
    "name": "OGP",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "-19021666:",
    "id": 10
},
{
    "domain": ".google.com",
    "expirationDate": 1610709687,
    "hostOnly": false,
    "httpOnly": false,
    "name": "OGPC",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "19021151-1:19021554-1:19021666-1:",
    "id": 11
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712543,
    "hostOnly": false,
    "httpOnly": false,
    "name": "SAPISID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "U7Jvl7_7Sluz0hBG/A6Am1zdaoueAHRNbq",
    "id": 12
},
{
    "domain": ".google.com",
    "expirationDate": 1622720442.963475,
    "hostOnly": false,
    "httpOnly": false,
    "name": "SEARCH_SAMESITE",
    "path": "/",
    "sameSite": "strict",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "CgQIqZEB",
    "id": 13
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712,
    "hostOnly": false,
    "httpOnly": false,
    "name": "SID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "4QdypGEtRQ9ATOo97TQkns3_Qm0FTjIQPwMwblSRZb1rAElh8gzgOALG1yiXPB386YSsvQ.",
    "id": 14
},
{
    "domain": ".google.com",
    "expirationDate": 1639807959.819353,
    "hostOnly": false,
    "httpOnly": false,
    "name": "SIDCC",
    "path": "/",
    "sameSite": "unspecified",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "AJi4QfHgC-2q553C-Y1Ihw9uT7lceyMctcypyvbJJzbqjaz92lA89USsVrtIhvDMz3l4PT4NIiQ",
    "id": 15
},
{
    "domain": ".google.com",
    "expirationDate": 1670238742.712409,
    "hostOnly": false,
    "httpOnly": true,
    "name": "SSID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": true,
    "session": false,
    "storeId": "0",
    "value": "APIyPDZX-wGspLDOD",
    "id": 16
}
]'''  ###cookies内容

cookies2 = json.loads(cookies2_txt)  ###cookies json化

driver.delete_all_cookies()  ###清除原有的cookies

for cook in cookies2:  ###cookies去sameSite开始

    try:

        cook.pop('sameSite')

    except:

        pass

    driver.add_cookie(cook)  ###cookies去sameSite结束，并将cookie添加入目标网址






while Count_Row < Max_Row + 1 :

    #driver.get(url)

    driver.get(url)  ###重新加载网址

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='gLFyf gsfi']")))  ###等待出现搜索店铺的搜索框加载完毕

    driver.find_element_by_xpath("//*[@class='gLFyf gsfi']").click()

    driver.find_element_by_xpath("//*[@class='gLFyf gsfi']").send_keys('澳门 facebook ',Restanrant_Search)

    driver.find_element_by_xpath("//*[@class='gLFyf gsfi']").send_keys(Keys.ENTER)

    time.sleep(1.1)




    driver.find_elements_by_xpath("//*[@class='LC20lb DKV0Md']")[0].click()

    time.sleep(2.2)

    Previous_Judgement_text = driver.current_url

    Previous_Judgement = Previous_Judgement_text.find('facebook')

    if Previous_Judgement != -1 :

        #Website_Judgement = driver.find_elements_by_xpath("//*[contains(text(),'www.')]")

        #print(Website_Judgement[0].text)

        #print(Website_Judgement[1].text)

        #print(Website_Judgement[2].text)

        #if len(Website_Judgement) == 0 :

            win32api.keybd_event(27, 0, 0, 0)

            win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

            Website = driver.current_url

            Sheet.cell(row=Count_Row, column=14).value = Website

            print(Website)

            print(Count_Row)

            Count_Row = Count_Row + 1

            Sheet.cell(row=1, column=31).value = Count_Row

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)



        #else :

            #win32api.keybd_event(27, 0, 0, 0)

            #win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

            #Website = Website_Judgement[1].text

            #Sheet.cell(row=Count_Row, column=14).value = Website

            #print(Website)

            #print(Count_Row)

            #Count_Row = Count_Row + 1

            #Sheet.cell(row=Count_Row, column=33).value = Count_Row

            #Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            #Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)




            #Origin_Windle_Handle = driver.current_window_handle

            #time.sleep(2.2)

            #driver.find_elements_by_xpath("//*[contains(text(),'www.')]")[2].click()

            #All_Window_Handles = driver.window_handles

            #for handle in All_Window_Handles:

                #if handle != Origin_Windle_Handle:

                    #driver.switch_to.window(handle)




            #time.sleep(2.5)

            #win32api.keybd_event(27, 0, 0, 0)

            #win32api.keybd_event(27, 0, win32con.KEYEVENTF_KEYUP, 0)

            #Website = driver.current_url

            #Booksite = driver.current_url

            #driver.close()

            #driver.switch_to.window(Origin_Windle_Handle)

            #Sheet.cell(row=Count_Row, column=14).value = Website

            #Sheet.cell(row=Count_Row, column=13).value = Booksite

            #print(Website)

            #print(Booksite)

            #print(Count_Row)

            #Count_Row = Count_Row + 1

            #Sheet.cell(row=Count_Row, column=31).value = Count_Row

            #Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)




    else :

        print(Count_Row)

        Count_Row = Count_Row + 1

        Sheet.cell(row=1, column=31).value = Count_Row

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)



print('搜寻完成')
print('搜寻完成')
print('搜寻完成')
print('搜寻完成')
print('搜寻完成')
print('搜寻完成')
print('搜寻完成')
###NT























