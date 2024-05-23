from selenium import webdriver
import time
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用


#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
option.add_argument("--headless")


experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)
#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表


######BUG修改区

#Count = 1

#Input_Row = 2

Count = 507       ###出错在这里修改

Input_Row = 508   ###出错在这里修改

######BUG修改区

Polld_Source = Sheet.cell(row=Input_Row, column=1).value




url = 'https://eo.dianping.com/epassport/bookauthmanage'

driver.get(url)

cookies1_txt = '''[
{
    "domain": ".dianping.com",
    "expirationDate": 1669845829,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_ga",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.723816111.1603966407",
    "id": 1
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606773889,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gat_gtag_UA_111805100_1",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 2
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606860229,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.544243607.1606512602",
    "id": 3
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635502415,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_hc.v",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "25c06052-ec24-b321-75e2-df2018908df6.1603966415",
    "id": 4
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 5
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_cuid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 6
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606775630,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_s",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1761b2ef3b3-bfe-0a3-97b%7C%7CNaN",
    "id": 7
},
{
    "domain": ".dianping.com",
    "expirationDate": 1638251852.03499,
    "hostOnly": false,
    "httpOnly": false,
    "name": "aburl",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 8
},
{
    "domain": ".dianping.com",
    "expirationDate": 1607268580.106885,
    "hostOnly": false,
    "httpOnly": false,
    "name": "bsid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
    "id": 9
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635761147.17086,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ctu",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "9f07eccced25cf82543a98c2d6e0235399e464f43265dd96bf67e9790d7d49ed",
    "id": 10
},
{
    "domain": ".dianping.com",
    "expirationDate": 1638251852.035075,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cy",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "2374",
    "id": 11
},
{
    "domain": ".dianping.com",
    "expirationDate": 1638251852.035129,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cye",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "osaka",
    "id": 12
},
{
    "domain": ".dianping.com",
    "expirationDate": 1607268580.107027,
    "hostOnly": false,
    "httpOnly": false,
    "name": "edper",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
    "id": 13
},
{
    "domain": ".dianping.com",
    "expirationDate": 1638251852,
    "hostOnly": false,
    "httpOnly": false,
    "name": "Hm_lvt_602b80cf8079ae6591966cc70a3940e7",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1604835193,1605286882,1605628795,1606715221",
    "id": 14
},
{
    "domain": ".dianping.com",
    "expirationDate": 1667994073.957559,
    "hostOnly": false,
    "httpOnly": false,
    "name": "s_ViewType",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "10",
    "id": 15
},
{
    "domain": ".dianping.com",
    "expirationDate": 1636371292.046143,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ua",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "dpuser_83084044340",
    "id": 16
},
{
    "domain": ".eo.dianping.com",
    "expirationDate": 1607378630,
    "hostOnly": false,
    "httpOnly": false,
    "name": "os-agent",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "",
    "id": 17
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "JSESSIONID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "0928A6A80AA5DB0CDB5AC92BB4C55DD0",
    "id": 18
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_custom_report",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "",
    "id": 19
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_session_token",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "goq7whd8fdy3h66du6qj",
    "id": 20
}
]'''          ###cookies内容

cookies1 = json.loads(cookies1_txt)###cookies json化

driver.delete_all_cookies()      ###清除原有的cookies

for cook in cookies1 :            ###cookies去sameSite开始
    try :
        cook.pop('sameSite')
    except :
        pass
    driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址
driver.get(url)                  ###重新加载网址




########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件








while Count < Max_Row :

    time.sleep(0.3)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[1]/ul/li[5]/div").click()                                                          ###找到并点击POLLD查询按钮

    time.sleep(0.3)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/form/div[1]/div[2]/div/span/input").clear()                          ###清除查询框已有文本

    time.sleep(0.3)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/form/div[1]/div[2]/div/span/input").send_keys(Polld_Source)            ###找到查询框并输入相应餐厅的POLLD号码

    time.sleep(0.3)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/form/div[2]/div/div/span/button").click()                            ###找到并点击查询按钮

    time.sleep(0.3)         ###等待0.1秒

    store_Location = driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/div/div/div/div/div/table/tbody/tr/td[5]")      ###找到餐厅地址

    text_Sourse1 = store_Location.text                                                      ###将餐厅地址文本化

    text_Sourse2 = re.sub("[A-Za-z0-9\!\%\[\]\,\.\，\。\-\〒\|]","",text_Sourse1)                   ###将餐厅地址只保留汉字

    text_Sourse3 = re.sub("[\u0800-\u4e00]+", "", text_Sourse2)                             ###将餐厅地址去除日文平片假名

    text_Sourse3 = re.sub("[\uff01-\uffee]+", "", text_Sourse3)

    text_Sourse3 = re.sub("[\u3040-\u309f]+", "", text_Sourse3)

    text_Sourse3 = re.sub("[\u30a0-\u30ff]+", "", text_Sourse3)

    text_Sourse3 = re.sub("[\u0100-\u024F]+", "", text_Sourse3)

    text_Sourse3 = text_Sourse3.replace(' ','')                                              ###将餐厅地址去空格

    text_Output = text_Sourse3.replace('　','')                                              ###将餐厅地址去空格

    List_text_Output = list(text_Output)





    if len(List_text_Output) != 0:

        print(text_Output)

        print(Count)

        Sheet.cell(row =Input_Row,column = 8).value = text_Output                                      ###向规定单元格写入数据

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

    if len(List_text_Output) == 0:

        text_Output = '####'

        print(text_Output)

        print(Count)

        Sheet.cell(row =Input_Row,column = 8).value = text_Output                                      ###向规定单元格写入数据

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


    Count = Count + 1

    Input_Row = Input_Row + 1

    Polld_Source = Sheet.cell(row=Input_Row, column=1).value




time.sleep(0.3)

print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")
print("已全部完毕。")




########除空插件

Count_Row = 1

None_Count = Sheet.cell(row=Count_Row, column=8).value

while Count_Row < Max_Row + 1 :

    if None_Count == '####' :

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row

        None_Count = Sheet.cell(row=Count_Row, column=8).value




        Total_Row = 1

        Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        while Total_Quantity != None:

            Total_Row = Total_Row + 1

            Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        Max_Row = Total_Row - 1

        print(Max_Row)




    if None_Count != '####' :

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=8).value

time.sleep(0.3)

print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")


########除空插件###NT


