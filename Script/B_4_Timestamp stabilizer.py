from nltk.tokenize import word_tokenize
import re
from openpyxl import load_workbook          ###Excel写入用
import time



Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表








########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件




Count_Row = 2               ###出错在这里修改

text = Sheet.cell(row=Count_Row, column=5).value








while Count_Row < Max_Row + 1 :

    Bigmodel = re.compile(u'[\u0065-\u0090]')

    Time_Judgement_1 = Bigmodel.search(text)

    Smallmodel = re.compile(u'[\u0097-\u0122]')

    Time_Judgement_2 = Smallmodel.search(text)

    if Time_Judgement_1 or Time_Judgement_2 :

        Sheet.cell(row=Count_Row, column=15).value = '17'

        Sheet.cell(row=Count_Row, column=16).value = '00'

        Sheet.cell(row=Count_Row, column=17).value = '20'

        Sheet.cell(row=Count_Row, column=18).value = '30'

        print('17 ','00 ','20 ','30')

        print(Count_Row)

        Sheet.cell(row=Count_Row, column=4).value = '17:00-20:30'

        Count_Row = Count_Row + 1

        text = Sheet.cell(row=Count_Row, column=5).value

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


    else :

        text = text.replace(' - ','-')

        text = text.replace(' ー ','-')

        text = text.replace('ー','-')

        text = text.replace('"','')

        #text = re.sub("[\!\%\,\〒\,\，\。\(\)\（\）\［\］\"]", "", text)

        #text = re.sub("[\u0065-\u0090]+", "", text)

        #text = re.sub("[\u0097-\u0122]+", "", text)

        #text = re.sub("[\u0800-\u9fa5]+", " ", text)

        #print(text)

        New_List_Source = word_tokenize(text,"english")

        print(Count_Row)

        print('时间结果 ',New_List_Source)

        print('len长 ',len(New_List_Source))


        if len(New_List_Source) == 0:

            Sheet.cell(row=Count_Row, column=15).value = '####'

            Count_Row = Count_Row + 1

            text = Sheet.cell(row=Count_Row, column=5).value

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


        if len(New_List_Source) == 1 :

            Time_1 = New_List_Source[0]

            if Time_1.count('pm') == 0 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

            if Time_1.count('pm') == 1 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_1[2] = str(int(Time_1[2]) + int(12))

            if Time_1.count('pm') == 2:

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_1[0] = str(int(Time_1[0]) + int(12))

                Time_1[2] = str(int(Time_1[2]) + int(12))

            if Time_1[1] == '15' :

                Time_1[1] = '00'

            if Time_1[1] == '45' :

                Time_1[1] = '30'

            if Time_1[3] == '15' :

                Time_1[3] = '00'

            if Time_1[3] == '45' :

                Time_1[3] = '30'

            if int(Time_1[2]) - int(Time_1[0]) < 0 and 24 - int(Time_1[2]) >= 0:

                Time_1[2] = 23

                Time_1[3] = 30

            if int(Time_1[2]) >= 24 :

                Time_1[2] = 23

            if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                Sheet.cell(row=Count_Row, column=17).value = 23

                Sheet.cell(row=Count_Row, column=18).value = 30

                Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                Count_Row = Count_Row + 1

                text = Sheet.cell(row=Count_Row, column=5).value

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




            else :

                Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                Count_Row = Count_Row + 1

                text = Sheet.cell(row=Count_Row, column=5).value

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




        if len(New_List_Source) > 1 :

            Time_1 = New_List_Source[0]

            Time_2 = New_List_Source[1]

            if Time_1.count('pm') == 0 and Time_2.count('pm') == 0:

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

            if Time_1.count('pm') == 1 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

                Time_1[2] = str(int(Time_1[2]) + int(12))

                Time_2[0] = str(int(Time_2[0]) + int(12))

                Time_2[2] = str(int(Time_2[2]) + int(12))

            if Time_1.count('pm') == 0 and Time_2.count('pm') > 0 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

                Time_2[0] = str(int(Time_2[0]) + int(12))

                Time_2[2] = str(int(Time_2[2]) + int(12))




            if Time_1[1] == '15' :

                Time_1[1] = '00'

            if Time_1[1] == '45' :

                Time_1[1] = '30'

            if Time_1[3] == '15' :

                Time_1[3] = '00'

            if Time_1[3] == '45' :

                Time_1[3] = '30'

            if Time_2[1] == '15' :

                Time_2[1] = '00'

            if Time_2[1] == '45' :

                Time_2[1] = '30'

            if Time_2[3] == '15' :

                Time_2[3] = '00'

            if Time_2[3] == '45' :

                Time_2[3] = '30'

            if int(Time_1[2]) - int(Time_1[0]) <= 0 :

                Time_1[2] = '23'




            if int(Time_2[0]) - int(Time_1[2]) >= 0 :

                if int(Time_2[2]) - int(Time_2[0]) < 0 and 24 - int(Time_2[2]) >= 0:

                    Time_2[2] = 23

                    Time_2[3] = 30

                if int(Time_2[2]) >= 24 :

                    Time_2[2] = 23

                if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=19).value = Time_2[0]

                    Sheet.cell(row=Count_Row, column=20).value = Time_2[1]

                    Sheet.cell(row=Count_Row, column=21).value = 23

                    Sheet.cell(row=Count_Row, column=22).value = 30

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0]) + '  ' + str(New_List_Source[1])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




                else:

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=19).value = Time_2[0]

                    Sheet.cell(row=Count_Row, column=20).value = Time_2[1]

                    Sheet.cell(row=Count_Row, column=21).value = str(int(Time_2[2]) - 1)

                    Sheet.cell(row=Count_Row, column=22).value = Time_2[3]

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0]) + '  ' + str(New_List_Source[1])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            if int(Time_2[0]) - int(Time_1[2]) < 0 :

                if int(Time_1[2]) - int(Time_1[0]) < 0 and 24 - int(Time_1[2]) >= 0 :

                    Time_1[2] = 23

                    Time_1[3] = 30

                if int(Time_1[2]) >= 24 :

                    Time_1[2] = 23

                if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = 23

                    Sheet.cell(row=Count_Row, column=18).value = 30

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




                else:

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


########除空插件

Count_Row = 1

None_Count = Sheet.cell(row=Count_Row, column=15).value

while Count_Row < Max_Row + 1 :

    if None_Count == '####' :

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row

        None_Count = Sheet.cell(row=Count_Row, column=15).value




        Total_Row_Sheet = 1

        Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        while Total_Quantity != None:

            Total_Row_Sheet = Total_Row_Sheet + 1

            Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        Max_Row_Sheet = Total_Row_Sheet - 1

        print(Max_Row_Sheet)




    if None_Count != '####' :

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=15).value

time.sleep(0.3)

print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")


########除空插件###NT










