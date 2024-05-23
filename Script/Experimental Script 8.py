from selenium import webdriver
import time
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用




Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表

######BUG修改区

#Count = 1

#Input_Row = 2

Count = 1

Input_Row = 2

######BUG修改区

Polld_Source = Sheet.cell(row=Input_Row, column=1).value




Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1

print(Max_Row)




Count_Row = 1

None_Count = Sheet.cell(row=Count_Row, column=8).value

while Count_Row < Max_Row + 1 :

    if None_Count == '####':

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=8).value




        Total_Row = 1

        Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        while Total_Quantity != None:

            Total_Row = Total_Row + 1

            Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        Max_Row = Total_Row - 1

        print(Max_Row)




    if None_Count != None :

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=8).value

time.sleep(0.3)

print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
###NT



































