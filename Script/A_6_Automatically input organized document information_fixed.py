from selenium import webdriver
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import nltk                                 ###自然语言处理用
import time
from selenium.webdriver.common.action_chains import ActionChains






#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
#option.add_argument("--headless")


experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)
#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表




Count = 1

Output_Row = 2

Output_Column_2_1_1 = 19

Output_Column_Week = 26

Polld_Source = Sheet.cell(row=Output_Row, column=1).value

Single_OR_Double = Sheet.cell(row=Output_Row, column=Output_Column_2_1_1).value

Week_Source = Sheet.cell(row=Output_Row, column=26).value

Telephone_Source = Sheet.cell(row=Output_Row, column=12).value


########Excel计数插件A

Total_Row_Sheet = 1

Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

while Total_Quantity != None :

    Total_Row_Sheet = Total_Row_Sheet + 1

    Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

Max_Row_Sheet = Total_Row_Sheet - 1

print(Max_Row_Sheet)

########Excel计数插件A











url = 'https://eo.dianping.com/epassport/bookauthmanage'

driver.get(url)

cookies1_txt = '''[
{
    "domain": ".dianping.com",
    "expirationDate": 1669735862,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_ga",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.723816111.1603966407",
    "id": 1
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606663922,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gat_gtag_UA_111805100_1",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 2
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606750262,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.544243607.1606512602",
    "id": 3
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635502415,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_hc.v",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "25c06052-ec24-b321-75e2-df2018908df6.1603966415",
    "id": 4
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 5
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_cuid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 6
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606665662,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_s",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "176149fb99b-60e-a06-449%7C%7C1",
    "id": 7
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094172,
    "hostOnly": false,
    "httpOnly": false,
    "name": "aburl",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 8
},
{
    "domain": ".dianping.com",
    "expirationDate": 1607268580.106885,
    "hostOnly": false,
    "httpOnly": false,
    "name": "bsid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
    "id": 9
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635761147.17086,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ctu",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "9f07eccced25cf82543a98c2d6e0235399e464f43265dd96bf67e9790d7d49ed",
    "id": 10
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094249,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cy",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "2374",
    "id": 11
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094289,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cye",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "osaka",
    "id": 12
},
{
    "domain": ".dianping.com",
    "expirationDate": 1607268580.107027,
    "hostOnly": false,
    "httpOnly": false,
    "name": "edper",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
    "id": 13
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659,
    "hostOnly": false,
    "httpOnly": false,
    "name": "Hm_lvt_602b80cf8079ae6591966cc70a3940e7",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1604647683,1604835193,1605286882,1605628795",
    "id": 14
},
{
    "domain": ".dianping.com",
    "expirationDate": 1667994073.957559,
    "hostOnly": false,
    "httpOnly": false,
    "name": "s_ViewType",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "10",
    "id": 15
},
{
    "domain": ".dianping.com",
    "expirationDate": 1636371292.046143,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ua",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "dpuser_83084044340",
    "id": 16
},
{
    "domain": ".eo.dianping.com",
    "expirationDate": 1607268662,
    "hostOnly": false,
    "httpOnly": false,
    "name": "os-agent",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "",
    "id": 17
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "JSESSIONID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "1715B488C3EDE3BE27F67B2C5782285E",
    "id": 18
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_custom_report",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "",
    "id": 19
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_session_token",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "jr2ykj71td5ykdn3ngch",
    "id": 20
}
]'''          ###cookies内容

cookies1 = json.loads(cookies1_txt)###cookies json化

driver.delete_all_cookies()      ###清除原有的cookies

for cook in cookies1 :            ###cookies去sameSite开始
    try :
        cook.pop('sameSite')
    except :
        pass
    driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址

#time.sleep(2)

driver.get(url)                  ###重新加载网址







while Output_Row < Max_Row_Sheet + 1 :               ###这里将来要改循环

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[1]/div[1]/div/div[2]/div/span/input")))  ###等待输入餐厅的POLLD号码框加载完毕

    time.sleep(1)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[1]/div[1]/div/div[2]/div/span/input").send_keys(Polld_Source)                 ###输入餐厅的POLLD号码

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[3]/div/button")))  ###等待输入餐厅的POLLD号码框加载完毕

    time.sleep(1)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[3]/div/button").click()                                                     ###点击查询

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/table/tbody/tr/td[8]/div/span[1]/a")))  ###等待输入餐厅的POLLD号码框加载完毕

    time.sleep(1)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/table/tbody/tr/td[8]/div/span[1]/a").click()                 ###加载完毕后，点击修改授权按钮
                                 #//*[@id="app"]/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/table/tbody/tr/td[8]/div/span[1]/a
    time.sleep(1)

    Handle_1 = driver.current_url




    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[2]/div[2]/div/span/div/div[4]/input")))  ###等待输入餐厅的POLLD号码框加载完毕

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[2]/div[2]/div/span/div/div[4]/input").clear()

    time.sleep(1)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[2]/div[2]/div/span/div/div[4]/input").send_keys(Telephone_Source)              ###输入餐厅联系电话






    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[9]/button[2]")))  ###等待输入餐厅的POLLD号码框加载完毕

    time.sleep(0.5)

    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[9]/button[2]").click()                                                      ###点击’提交授权‘

                                 #//*[@id="app"]/div/div/div[2]/div[2]/div/div/form[2]/div[9]/button[2]






    time.sleep(1.5)

    print(Output_Row)

    Count = Count + 1

    Output_Row = Output_Row + 1

    Polld_Source = Sheet.cell(row=Output_Row, column=1).value

    Single_OR_Double = Sheet.cell(row=Output_Row, column=Output_Column_2_1_1).value

    Week_Source = Sheet.cell(row=Output_Row, column=26).value

    Telephone_Source = Sheet.cell(row=Output_Row, column=12).value



    #Handle_2 = driver.current_url




    if Output_Row < Max_Row_Sheet + 1 :            ###这里将来要改循环

        time.sleep(0.7)

        driver.execute_script("window.open('{}');".format(url))

        time.sleep(0.7)

        driver.close()

        time.sleep(0.7)

        driver.switch_to.window(driver.window_handles[-1])

        time.sleep(0.7)




#Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

driver.quit()

print('录入完成')
print('录入完成')
print('录入完成')
print('录入完成')
print('录入完成')
print('录入完成')
print('录入完成')
print('录入完成')
###NT












