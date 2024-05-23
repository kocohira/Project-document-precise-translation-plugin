import re  ###正则表达式用
from openpyxl import load_workbook  ###Excel写入用

Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')  ###读取Excel比较表格

Sheet = Excel['澳门']  ###打开‘大阪’工作表

Count_Row = 2  ###每次出错就从这里修改

########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None:
    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1  ###包含了表格第一行

print(Max_Row)

########Excel计数插件




########处理预定方式

#Count_Row = 2

#while Count_Row < Max_Row + 1:

    #Booksite = Sheet.cell(row=Count_Row, column=13).value

    #if Booksite != None:

        #Sheet.cell(row=Count_Row, column=17).value = '官网'

        #print(Count_Row)

        #Excel.save(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')

        #Count_Row = Count_Row + 1

    #else:

        #Sheet.cell(row=Count_Row, column=17).value = '电话'

        #print(Count_Row)

        #Excel.save(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')

        #Count_Row = Count_Row + 1

print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')

########处理预定方式


########处理是否授权

Count_Row = 86

while Count_Row < Max_Row + 1:

    Authorization = float(Sheet.cell(row=Count_Row, column=3).value)

    if Authorization <= 150:

        Sheet.cell(row=Count_Row, column=18).value = '已设置授权'

        print(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')

        Count_Row = Count_Row + 1

    else:

        Sheet.cell(row=Count_Row, column=18).value = '可授权'

        print(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')

        Count_Row = Count_Row + 1

print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')

########处理是否授权


########署名

Count_Row = 2

while Count_Row < Max_Row + 1:
    Sheet.cell(row=Count_Row, column=21).value = 'Jack'

    print(Count_Row)

    Excel.save(r'C:\Users\kocohira\Desktop\Special\澳门绿体字三期—Jack.xlsx')

    Count_Row = Count_Row + 1

print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')
print('处理结束')

########署名###NT


