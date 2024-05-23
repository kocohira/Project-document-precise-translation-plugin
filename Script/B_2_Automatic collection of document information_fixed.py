

from selenium import webdriver
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import nltk                                 ###自然语言处理用
from nltk.tokenize import word_tokenize
import time,win32api,win32con
import os
import sys



def restart_program() :

    print('重启程序中...')

    python = sys.executable

    os.execl(python,python,*sys.argv)




#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
#option.add_argument("--headless")

experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)



#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

Sheet = Excel['大阪']  ###打开‘大阪’工作表

Count_Row = int((Sheet.cell(row=1, column=31)).value)                             ###每次出错就从这里修改

Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)                ###搜索框中要输入的店铺

Location_Search = 'Macau'                 ###搜索框中要输入的区域




########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件








url = 'https://www.openrice.com/en/macau/r-u-cafe-veng-kei-flores-international-salad-r486723'



driver.get(url)




while Count_Row < Max_Row + 1 :

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='header-searchbar-where-input']")))  ###等待出现搜索店铺的搜索框加载完毕

    driver.find_element_by_xpath("//*[@id='header-searchbar-where-input']").click()

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='where quick-search-options js-quick-search-options or-scrollbar']")))  ###等待出现搜索店铺的搜索框加载完毕

    driver.find_element_by_xpath("//*[@class='where quick-search-options js-quick-search-options or-scrollbar']/dl/dt/dl/dt/div").click()

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='header-searchbar-what-input what js-input-quick-search']")))  ###等待出现搜索店铺的搜索框加载完毕

    #driver.find_element_by_xpath("//*[@class='header-searchbar-what-input what js-input-quick-search']").click()

    driver.find_element_by_xpath("//*[@class='header-searchbar-what-input what js-input-quick-search']").send_keys(Restanrant_Search)

    driver.find_element_by_xpath("//*[@class='btn header-searchbar-search-btn']").click()




    time.sleep(2.5)

    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='js-poi-list-main']")))  ###等待出现搜索店铺的搜索框加载完毕

    Restanrant_Quantity = driver.find_elements_by_xpath("//*[text()='No results were found.']")

    Restanrant_Judgement = len(Restanrant_Quantity)

    if Restanrant_Judgement == 0 :

        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='sr1-listing-content-cells pois-restaurant-list js-poi-list-content-cell-container']/li/div/section/div/h2/a")))  ###等待出现搜索店铺的搜索框加载完毕

        driver.find_element_by_xpath("//*[@class='sr1-listing-content-cells pois-restaurant-list js-poi-list-content-cell-container']/li/div/section/div/h2/a").click()

        time.sleep(1.5)

        #windows = driver.window_handles                                                                     ####搜索当前所有打开的页面的句柄（识别用）

        #driver.close()                                                                                      ###关闭上一个页面

        #driver.switch_to.window(windows[-1])






        ###在这里判断店铺有无电话，地址，营业时间

        Location_Judgement_1 = driver.find_elements_by_xpath("//*[@id='global-container']/main/div[2]/div[2]/div[1]/div/div[1]/section/div[1]/div[1]/div[1]")

        if len(Location_Judgement_1) >= 1 :

            Location_Source = driver.find_element_by_xpath("//*[@class='left-col']/div/div/section/div[2]/a").text

            zhmodel = re.compile(u'[\u4e00-\u9fa5]')

            Location_Judgement_2 = zhmodel.search(Location_Source)

            if Location_Judgement_2 :         ###说不定要改

                Telephone_Judgement_1 = driver.find_elements_by_xpath("//*[@id='global-container']/main/div[2]/div[2]/div[1]/div/div[2]/section/div[1]")

                if len(Telephone_Judgement_1) >= 1 :

                    Telephone_Source = driver.find_element_by_xpath("//*[@class='telephone-section']/div[2]").text

                    numbermodel = re.compile('[0-9]+')

                    Telephone_Judgement_2 = numbermodel.findall(Telephone_Source)

                    if Telephone_Judgement_2 :

                        Telephone_Source = Telephone_Source[5:]

                        Telephone_Source = Telephone_Source.replace(' ','')

                        Time_Judgement = driver.find_elements_by_xpath("//*[text()='Opening Hours']")

                        if len(Time_Judgement) >= 1 :

                            Openrice_Time_source = driver.find_elements_by_xpath("//*[@class='opening-hours-time']")[1].text

                            Week_Judgement = driver.find_elements_by_xpath("//*[@class='opening-hours-day']")

                            if len(Week_Judgement) == 2 :

                                Openrice_Week = '0'

                            if len(Week_Judgement) > 2 :

                                Week_Judgement_Sub_1_Source = driver.find_element_by_xpath("//*[@class='opening-hours-section js-normal-and-special-opening-hours-section']/div/div/div").text

                                Week_Judgement_Sub_2_Source = driver.find_element_by_xpath("//*[@class='opening-hours-section js-normal-and-special-opening-hours-section']/div/div[2]/div").text

                                Week_Judgement_Sub_1_list = word_tokenize(Week_Judgement_Sub_1_Source, "english")

                                Week_Judgement_Sub_2_list = word_tokenize(Week_Judgement_Sub_2_Source, "english")

                                if len(Week_Judgement_Sub_1_list) == 1 :

                                    if Week_Judgement_Sub_1_list[0] == 'Tue' :

                                        Openrice_Week = '1'

                                    elif Week_Judgement_Sub_1_list[0] == 'Wed' :

                                        Openrice_Week = '2'

                                    elif Week_Judgement_Sub_1_list[0] == 'Thu' :

                                        Openrice_Week = '3'

                                    else :

                                        Openrice_Week = '0'




                                elif len(Week_Judgement_Sub_1_list) == 3 :

                                    Week_Judgement_Sub_1_2 = Week_Judgement_Sub_1_list[2]

                                    Week_Judgement_Sub_2_1 = Week_Judgement_Sub_2_list[0]

                                    if len(Week_Judgement_Sub_2_list) == 3 :

                                        if Week_Judgement_Sub_1_2 == 'Tue' and Week_Judgement_Sub_2_1 == 'Thu' :

                                            Openrice_Week = '3'

                                        if Week_Judgement_Sub_1_2 == 'Wed' and Week_Judgement_Sub_2_1 == 'Fri':

                                            Openrice_Week = '4'

                                        if Week_Judgement_Sub_1_2 == 'Thu' and Week_Judgement_Sub_2_1 == 'Sat':

                                            Openrice_Week = '5'

                                        if Week_Judgement_Sub_1_2 == 'Fri' and Week_Judgement_Sub_2_1 == 'Sun':

                                            Openrice_Week = '6'

                                        else :

                                            Openrice_Week = '0'

                                    if len(Week_Judgement_Sub_2_list) == 1 :

                                        if Week_Judgement_Sub_2_1 == 'Sat' :

                                            Openrice_Week = '7'

                                        if Week_Judgement_Sub_2_1 == 'Fri' :

                                            Openrice_Week = '8'

                                        else :

                                            Openrice_Week = '0'




                                else :

                                    Openrice_Week = '0'
#



                            Price_Judgement_1 = driver.find_elements_by_xpath("//*[@class='header-poi-price dot-separator']/a")

                            if len(Price_Judgement_1) >= 1 :

                                Price_Judgement_2_Source = driver.find_element_by_xpath("//*[@class='header-poi-price dot-separator']/a").text

                                Price_Judgement_2_Source = Price_Judgement_2_Source.replace(' ','')

                                Price_Judgement_2_Source = re.sub("[\$]", "", Price_Judgement_2_Source)

                                Price_Judgement_2_Source = re.sub("[A-Za-z]", "", Price_Judgement_2_Source)

                                Price_Judgement_2_Source = Price_Judgement_2_Source.replace('-', ' ')

                                Price_Judgement_2 = word_tokenize(Price_Judgement_2_Source, "english")

                                if len(Price_Judgement_2) == 1 :

                                    Openrice_Price = '9.9'

                                if len(Price_Judgement_2) == 2 :

                                    if int(Price_Judgement_2[0]) >= 150 :

                                        Openrice_Price = str(int(Price_Judgement_2[0]))

                                    else :

                                        if int(Price_Judgement_2[1]) < 150 :

                                            Openrice_Price = '9.9'

                                        if int(Price_Judgement_2[1]) >= 150 :

                                            Openrice_Price = str(int(Price_Judgement_2[1]))




                            print(Openrice_Price)

                            print(Openrice_Time_source)

                            print(Location_Source)

                            print(Telephone_Source)

                            Sheet.cell(row=Count_Row, column=3).value = Openrice_Price

                            Sheet.cell(row=Count_Row, column=5).value = Openrice_Time_source

                            Sheet.cell(row=Count_Row, column=11).value = Location_Source

                            Sheet.cell(row=Count_Row, column=12).value = Telephone_Source

                            Sheet.cell(row=Count_Row, column=26).value = Openrice_Week

                            print (Count_Row)

                            Count_Row = Count_Row + 1

                            Sheet.cell(row=1, column=31).value = Count_Row

                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                            driver.get(url)

                            time.sleep(2.5)

                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                            # driver.close()  ###关闭上一个页面

                            driver.switch_to.window(windows[-1])

                            time.sleep(2.5)


                        else :

                            Openrice_Time_source = '16:00-20:00'

                            Openrice_Week = '0'

                            Openrice_Price = '9.9'




                            print(Openrice_Price)

                            print(Openrice_Time_source)

                            print(Location_Source)

                            print(Telephone_Source)

                            Sheet.cell(row=Count_Row, column=3).value = Openrice_Price

                            Sheet.cell(row=Count_Row, column=5).value = Openrice_Time_source

                            Sheet.cell(row=Count_Row, column=11).value = Location_Source

                            Sheet.cell(row=Count_Row, column=12).value = Telephone_Source

                            Sheet.cell(row=Count_Row, column=26).value = Openrice_Week

                            print (Count_Row)

                            Count_Row = Count_Row + 1

                            Sheet.cell(row=1, column=31).value = Count_Row

                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                            driver.get(url)

                            time.sleep(2.5)

                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                            # driver.close()  ###关闭上一个页面

                            driver.switch_to.window(windows[-1])

                            time.sleep(2.5)





                    else :

                        print(Count_Row)

                        Sheet.cell(row=Count_Row, column=12).value = '####'

                        Count_Row = Count_Row + 1

                        Sheet.cell(row=1, column=31).value = Count_Row

                        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                        driver.get(url)

                        time.sleep(2.5)

                        windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                        # driver.close()  ###关闭上一个页面

                        driver.switch_to.window(windows[-1])

                        time.sleep(2.5)




                else :

                    print(Count_Row)

                    Sheet.cell(row=Count_Row, column=12).value = '####'

                    Count_Row = Count_Row + 1

                    Sheet.cell(row=1, column=31).value = Count_Row

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                    driver.get(url)

                    time.sleep(2.5)

                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                    # driver.close()  ###关闭上一个页面

                    driver.switch_to.window(windows[-1])

                    time.sleep(2.5)




            else :

                print(Count_Row)

                Sheet.cell(row=Count_Row, column=11).value = '####'

                Count_Row = Count_Row + 1

                Sheet.cell(row=1, column=31).value = Count_Row

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                driver.get(url)

                time.sleep(2.5)

                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                # driver.close()  ###关闭上一个页面

                driver.switch_to.window(windows[-1])

                time.sleep(2.5)




        else :

            print(Count_Row)

            Sheet.cell(row=Count_Row, column=11).value = '####'

            Count_Row = Count_Row + 1

            Sheet.cell(row=1, column=31).value = Count_Row

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

            driver.get(url)

            time.sleep(2.5)

            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

            # driver.close()  ###关闭上一个页面

            driver.switch_to.window(windows[-1])

            time.sleep(2.5)




    else :

        print(Count_Row)

        Sheet.cell(row=Count_Row, column=11).value = '####'

        Count_Row = Count_Row + 1

        Sheet.cell(row=1, column=31).value = Count_Row

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

        driver.get(url)

        time.sleep(2.5)

        windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

        # driver.close()  ###关闭上一个页面

        driver.switch_to.window(windows[-1])

        time.sleep(2.5)



###NT


