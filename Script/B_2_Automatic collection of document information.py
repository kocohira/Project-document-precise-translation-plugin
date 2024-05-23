#! /usr/bin/env python3
#! /usr/bin/env python
import goto
from dominate.tags import label
from goto import with_goto

@with_goto
def Total() :

    label.begin

    try :

        from selenium import webdriver
        import json
        import re       ###正则表达式用
        from openpyxl import load_workbook          ###Excel写入用
        from datetime import datetime
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import nltk                                 ###自然语言处理用
        from nltk.tokenize import word_tokenize
        import time,win32api,win32con
        import os
        import sys



        def restart_program() :

            print('重启程序中...')

            python = sys.executable

            os.execl(python,python,*sys.argv)




        #实例化谷歌设置选项
        option = webdriver.ChromeOptions()
        #添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
        #option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
        option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
        #option.add_argument("--headless")

        experimentalFlags = [
            #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
            'same-site-by-default-cookies@2',
            #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
            'cookies-without-same-site-must-be-secure@2',]
        chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
        option.add_experimental_option('localState', chromeLocalStatePrefs)



        #初始化driver
        driver = webdriver.Chrome(options=option)








        Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

        Sheet = Excel['大阪']  ###打开‘大阪’工作表

        Count_Row = int((Sheet.cell(row=1, column=31)).value)                             ###每次出错就从这里修改

        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)                ###搜索框中要输入的店铺

        Location_Search = 'Macau'                 ###搜索框中要输入的区域




        ########Excel计数插件

        Total_Row = 1

        Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        while Total_Quantity != None :

            Total_Row = Total_Row + 1

            Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

        Max_Row = Total_Row - 1             ###包含了表格第一行

        print(Max_Row)

        ########Excel计数插件








        url = 'https://www.openrice.com/en/macau/r-u-cafe-veng-kei-flores-international-salad-r486723'



        driver.get(url)

        cookies2_txt =  '''[
        {
            "domain": ".openrice.com",
            "expirationDate": 1637670463,
            "hostOnly": false,
            "httpOnly": false,
            "name": "__gads",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "ID=36aac194d15ea58e:T=1603974463:S=ALNI_MYJBV51SahHmD_9YUr7_aakFECVyw",
            "id": 1
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1670755468,
            "hostOnly": false,
            "httpOnly": false,
            "name": "__utma",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "183676536.930136130.1603974463.1607208628.1607683469.17",
            "id": 2
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1607685268,
            "hostOnly": false,
            "httpOnly": false,
            "name": "__utmb",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "183676536.1.10.1607683469",
            "id": 3
        },
        {
            "domain": ".openrice.com",
            "hostOnly": false,
            "httpOnly": false,
            "name": "__utmc",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "183676536",
            "id": 4
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1607684068,
            "hostOnly": false,
            "httpOnly": false,
            "name": "__utmt_UA-652541-1",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 5
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1623451468,
            "hostOnly": false,
            "httpOnly": false,
            "name": "__utmz",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "183676536.1603984105.5.4.utmcsr=google|utmccn=(organic)|utmcmd=organic|utmctr=(not%20provided)",
            "id": 6
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1613839541,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_fbp",
            "path": "/",
            "sameSite": "lax",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "fb.1.1603974480744.264401216",
            "id": 7
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1670755473,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_ga",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.1.930136130.1603974463",
            "id": 8
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1670755473,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_ga_WM2DLEGHYP",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GS1.1.1607683473.16.0.1607683473.0",
            "id": 9
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1607769868,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_gid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.2.34561546.1607683469",
            "id": 10
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 2147385600,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_uid44664",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "DC4E2E7C.1",
            "id": 11
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1639219453.769577,
            "hostOnly": false,
            "httpOnly": false,
            "name": "autha",
            "path": "/",
            "sameSite": "unspecified",
            "secure": true,
            "session": false,
            "storeId": "0",
            "value": "INN5hFaDCRHABqILlwQmx3z1tQ_vxN10MgoQ6eFkNpBxg5Z02ZcaoUz2l8JvwIQscHxUrIBXYL7dbClyelYvyd7UkWVsvfl60Yetj8he5e64lUE2M3GSieqfJjFQw-3zJeY1kyroJSx5m7raFWgMKRVVjBc2cQDk3hthImguwsGLqpjSiVgSysoECnss4FQzGp9a1519GKcrTTJY3L7Q505davXyMseom4UzNQGLA300s3-9mhFyOHhpdsZMToPnhrHvhOcrCbpwI1aZTov8ssnOOx9P1dTsQldULRlAD46ikq6Fi8BAVwVyAMJ_UZR5aN7A-GKkomck_k9iC9Z2Mg7MdPrsNbAHEydHfD6IF07Tbp0NtEZ0djhZ9RxCPKUylI4tNsUK05icV8xwRu5jRlptsuuRYLRMriGJX01D7LX2Xv72PkZ3_mmmEMULGirZaZpkIQ",
            "id": 12
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1639219453.769713,
            "hostOnly": false,
            "httpOnly": false,
            "name": "authe",
            "path": "/",
            "sameSite": "unspecified",
            "secure": true,
            "session": false,
            "storeId": "0",
            "value": "aUKrbvhEpW8P1aIA1fp2MV81Pb6qHbL4nd1Qz5XrqNMTNMVWeksfID//QI5386cWtkioniFxvD4lVLS/1kULEwe3NFA8mipJY+gVMgthe7I=",
            "id": 13
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1639219453.769663,
            "hostOnly": false,
            "httpOnly": false,
            "name": "authr",
            "path": "/",
            "sameSite": "unspecified",
            "secure": true,
            "session": false,
            "storeId": "0",
            "value": "SJhIQcAgtoOGaMtrx74wERBevPC4PCemSdi07AE5fXeCFibzN_vDM3Y2tSuQRfHedmoZRZ9dC-MrV0MkiZ0GOBD51M75Zx2hJstq3cRG-bl_qi4eWfWzR5YBE8vf5OC1X5PTNbwPdGHBVgpS97V3RfZe6gz7e59HNpOrApFplkVH9reOlPS_mka0a3C64RWXZGE11gkkhclyYjzy7eQ2SWQ-VN8CtT-mStOxwS75J3ZdSwGbJiZYSfymCXTeupd6_An6AuniHzQV76y7BIT9cxUHZRnYYUM895YQB3SO-R3c1LVEJSpfp3_jnpiUFUs8TKP3Aw2cFXkQnoEKPjf-FcW0pu9WM3DDPdR6r7ea8ho9WNjDj1ZGnTFXFNHpe4ZaEjpOi-l5PUqSkjhEF8hqgqoFY-xhL5PPO6K04LDajmFYzCLW0nJIUb7S6SCpJF2jmJ0_R2ow77-e3-dsi6jSDAXENcQ",
            "id": 14
        },
        {
            "domain": ".openrice.com",
            "hostOnly": false,
            "httpOnly": false,
            "name": "DefaultRegionIds",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "%7B%22hk%22%3A1%7D",
            "id": 15
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1635510464.118145,
            "hostOnly": false,
            "httpOnly": false,
            "name": "isguest",
            "path": "/",
            "sameSite": "unspecified",
            "secure": true,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 16
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1667224729,
            "hostOnly": false,
            "httpOnly": false,
            "name": "iUUID",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "62691d1f09f72b15888c74ca22a35fe2",
            "id": 17
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1610304272,
            "hostOnly": false,
            "httpOnly": false,
            "name": "truvid_protected",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "{\"val\":\"c\",\"level\":2,\"geo\":\"CN\",\"timestamp\":1607683472}",
            "id": 18
        },
        {
            "domain": ".openrice.com",
            "expirationDate": 1919334463,
            "hostOnly": false,
            "httpOnly": false,
            "name": "webhash",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "e4d0fe08-4b7a-41b0-83fb-cff5c2a62759",
            "id": 19
        },
        {
            "domain": "www.openrice.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "GED_PLAYLIST_ACTIVITY",
            "path": "/",
            "sameSite": "unspecified",
            "secure": true,
            "session": true,
            "storeId": "0",
            "value": "W3sidSI6IjMxWFAiLCJ0c2wiOjE2MDc2ODM0NzUsIm52IjoxLCJ1cHQiOjE2MDc2ODM0NTQsImx0IjoxNjA3NjgzNDc0fV0.",
            "id": 20
        },
        {
            "domain": "www.openrice.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "RegionId",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "1",
            "id": 21
        },
        {
            "domain": "www.openrice.com",
            "expirationDate": 1638744721,
            "hostOnly": true,
            "httpOnly": false,
            "name": "ucfunnel_uid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "f45c45be-9ca6-3997-aaef-71db850b1906",
            "id": 22
        }
        ]'''         ###cookies内容
        '''
        cookies2 = json.loads(cookies2_txt)###cookies json化

        driver.delete_all_cookies()      ###清除原有的cookies

        for cook in cookies2 :            ###cookies去sameSite开始
            try :
                cook.pop('sameSite')
            except :
                pass
            driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址
        driver.get(url)                  ###重新加载网址
        '''



        while Count_Row < Max_Row + 1 :

            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='header-searchbar-where-input']")))  ###等待出现搜索店铺的搜索框加载完毕

            driver.find_element_by_xpath("//*[@id='header-searchbar-where-input']").click()

            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='where quick-search-options js-quick-search-options or-scrollbar']")))  ###等待出现搜索店铺的搜索框加载完毕

            driver.find_element_by_xpath("//*[@class='where quick-search-options js-quick-search-options or-scrollbar']/dl/dt/dl/dt/div").click()

            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='header-searchbar-what-input what js-input-quick-search']")))  ###等待出现搜索店铺的搜索框加载完毕

            #driver.find_element_by_xpath("//*[@class='header-searchbar-what-input what js-input-quick-search']").click()

            driver.find_element_by_xpath("//*[@class='header-searchbar-what-input what js-input-quick-search']").send_keys(Restanrant_Search)

            driver.find_element_by_xpath("//*[@class='btn header-searchbar-search-btn']").click()




            time.sleep(2.5)

            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='js-poi-list-main']")))  ###等待出现搜索店铺的搜索框加载完毕

            Restanrant_Quantity = driver.find_elements_by_xpath("//*[text()='No results were found.']")

            Restanrant_Judgement = len(Restanrant_Quantity)

            if Restanrant_Judgement == 0 :

                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@class='sr1-listing-content-cells pois-restaurant-list js-poi-list-content-cell-container']/li/div/section/div/h2/a")))  ###等待出现搜索店铺的搜索框加载完毕

                driver.find_element_by_xpath("//*[@class='sr1-listing-content-cells pois-restaurant-list js-poi-list-content-cell-container']/li/div/section/div/h2/a").click()

                time.sleep(1.5)

                #windows = driver.window_handles                                                                     ####搜索当前所有打开的页面的句柄（识别用）

                #driver.close()                                                                                      ###关闭上一个页面

                #driver.switch_to.window(windows[-1])






                ###在这里判断店铺有无电话，地址，营业时间

                Location_Judgement_1 = driver.find_elements_by_xpath("//*[@id='global-container']/main/div[2]/div[2]/div[1]/div/div[1]/section/div[1]/div[1]/div[1]")

                if len(Location_Judgement_1) >= 1 :

                    Location_Source = driver.find_element_by_xpath("//*[@class='left-col']/div/div/section/div[2]/a").text

                    zhmodel = re.compile(u'[\u4e00-\u9fa5]')

                    Location_Judgement_2 = zhmodel.search(Location_Source)

                    if Location_Judgement_2 :         ###说不定要改

                        Telephone_Judgement_1 = driver.find_elements_by_xpath("//*[@id='global-container']/main/div[2]/div[2]/div[1]/div/div[2]/section/div[1]")

                        if len(Telephone_Judgement_1) >= 1 :

                            Telephone_Source = driver.find_element_by_xpath("//*[@class='telephone-section']/div[2]").text

                            numbermodel = re.compile('[0-9]+')

                            Telephone_Judgement_2 = numbermodel.findall(Telephone_Source)

                            if Telephone_Judgement_2 :

                                Telephone_Source = Telephone_Source[5:]

                                Telephone_Source = Telephone_Source.replace(' ','')

                                Time_Judgement = driver.find_elements_by_xpath("//*[text()='Opening Hours']")

                                if len(Time_Judgement) >= 1 :

                                    Openrice_Time_source = driver.find_elements_by_xpath("//*[@class='opening-hours-time']")[1].text

                                    Week_Judgement = driver.find_elements_by_xpath("//*[@class='opening-hours-day']")

                                    if len(Week_Judgement) == 2 :

                                        Openrice_Week = '0'

                                    if len(Week_Judgement) > 2 :

                                        Week_Judgement_Sub_1_Source = driver.find_element_by_xpath("//*[@class='opening-hours-section js-normal-and-special-opening-hours-section']/div/div/div").text

                                        Week_Judgement_Sub_2_Source = driver.find_element_by_xpath("//*[@class='opening-hours-section js-normal-and-special-opening-hours-section']/div/div[2]/div").text

                                        Week_Judgement_Sub_1_list = word_tokenize(Week_Judgement_Sub_1_Source, "english")

                                        Week_Judgement_Sub_2_list = word_tokenize(Week_Judgement_Sub_2_Source, "english")

                                        if len(Week_Judgement_Sub_1_list) == 1 :

                                            if Week_Judgement_Sub_1_list[0] == 'Tue' :

                                                Openrice_Week = '1'

                                            elif Week_Judgement_Sub_1_list[0] == 'Wed' :

                                                Openrice_Week = '2'

                                            elif Week_Judgement_Sub_1_list[0] == 'Thu' :

                                                Openrice_Week = '3'

                                            else :

                                                Openrice_Week = '0'




                                        elif len(Week_Judgement_Sub_1_list) == 3 :

                                            Week_Judgement_Sub_1_2 = Week_Judgement_Sub_1_list[2]

                                            Week_Judgement_Sub_2_1 = Week_Judgement_Sub_2_list[0]

                                            if len(Week_Judgement_Sub_2_list) == 3 :

                                                if Week_Judgement_Sub_1_2 == 'Tue' and Week_Judgement_Sub_2_1 == 'Thu' :

                                                    Openrice_Week = '3'

                                                if Week_Judgement_Sub_1_2 == 'Wed' and Week_Judgement_Sub_2_1 == 'Fri':

                                                    Openrice_Week = '4'

                                                if Week_Judgement_Sub_1_2 == 'Thu' and Week_Judgement_Sub_2_1 == 'Sat':

                                                    Openrice_Week = '5'

                                                if Week_Judgement_Sub_1_2 == 'Fri' and Week_Judgement_Sub_2_1 == 'Sun':

                                                    Openrice_Week = '6'

                                                else :

                                                    Openrice_Week = '0'

                                            if len(Week_Judgement_Sub_2_list) == 1 :

                                                if Week_Judgement_Sub_2_1 == 'Sat' :

                                                    Openrice_Week = '7'

                                                if Week_Judgement_Sub_2_1 == 'Fri' :

                                                    Openrice_Week = '8'

                                                else :

                                                    Openrice_Week = '0'




                                        else :

                                            Openrice_Week = '0'
#



                                    Price_Judgement_1 = driver.find_elements_by_xpath("//*[@class='header-poi-price dot-separator']/a")

                                    if len(Price_Judgement_1) >= 1 :

                                        Price_Judgement_2_Source = driver.find_element_by_xpath("//*[@class='header-poi-price dot-separator']/a").text

                                        Price_Judgement_2_Source = Price_Judgement_2_Source.replace(' ','')

                                        Price_Judgement_2_Source = re.sub("[\$]", "", Price_Judgement_2_Source)

                                        Price_Judgement_2_Source = re.sub("[A-Za-z]", "", Price_Judgement_2_Source)

                                        Price_Judgement_2_Source = Price_Judgement_2_Source.replace('-', ' ')

                                        Price_Judgement_2 = word_tokenize(Price_Judgement_2_Source, "english")

                                        if len(Price_Judgement_2) == 1 :

                                            Openrice_Price = '9.9'

                                        if len(Price_Judgement_2) == 2 :

                                            if int(Price_Judgement_2[0]) >= 150 :

                                                Openrice_Price = str(int(Price_Judgement_2[0]))

                                            else :

                                                if int(Price_Judgement_2[1]) < 150 :

                                                    Openrice_Price = '9.9'

                                                if int(Price_Judgement_2[1]) >= 150 :

                                                    Openrice_Price = '9.9'




                                    print(Openrice_Price)

                                    print(Openrice_Time_source)

                                    print(Location_Source)

                                    print(Telephone_Source)

                                    Sheet.cell(row=Count_Row, column=3).value = Openrice_Price

                                    Sheet.cell(row=Count_Row, column=5).value = Openrice_Time_source

                                    Sheet.cell(row=Count_Row, column=11).value = Location_Source

                                    Sheet.cell(row=Count_Row, column=12).value = Telephone_Source

                                    Sheet.cell(row=Count_Row, column=26).value = Openrice_Week

                                    print (Count_Row)

                                    Count_Row = Count_Row + 1

                                    Sheet.cell(row=1, column=31).value = Count_Row

                                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                    driver.get(url)

                                    time.sleep(2.5)

                                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                    # driver.close()  ###关闭上一个页面

                                    driver.switch_to.window(windows[-1])

                                    time.sleep(2.5)


                                else :

                                    Openrice_Time_source = '16:00-20:00'

                                    Openrice_Week = '0'

                                    Openrice_Price = '9.9'




                                    print(Openrice_Price)

                                    print(Openrice_Time_source)

                                    print(Location_Source)

                                    print(Telephone_Source)

                                    Sheet.cell(row=Count_Row, column=3).value = Openrice_Price

                                    Sheet.cell(row=Count_Row, column=5).value = Openrice_Time_source

                                    Sheet.cell(row=Count_Row, column=11).value = Location_Source

                                    Sheet.cell(row=Count_Row, column=12).value = Telephone_Source

                                    Sheet.cell(row=Count_Row, column=26).value = Openrice_Week

                                    print (Count_Row)

                                    Count_Row = Count_Row + 1

                                    Sheet.cell(row=1, column=31).value = Count_Row

                                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                    driver.get(url)

                                    time.sleep(2.5)

                                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                    # driver.close()  ###关闭上一个页面

                                    driver.switch_to.window(windows[-1])

                                    time.sleep(2.5)





                            else :

                                print(Count_Row)

                                Sheet.cell(row=Count_Row, column=12).value = '####'

                                Count_Row = Count_Row + 1

                                Sheet.cell(row=1, column=31).value = Count_Row

                                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                                driver.get(url)

                                time.sleep(2.5)

                                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                                # driver.close()  ###关闭上一个页面

                                driver.switch_to.window(windows[-1])

                                time.sleep(2.5)




                        else :

                            print(Count_Row)

                            Sheet.cell(row=Count_Row, column=12).value = '####'

                            Count_Row = Count_Row + 1

                            Sheet.cell(row=1, column=31).value = Count_Row

                            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                            Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                            driver.get(url)

                            time.sleep(2.5)

                            windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                            # driver.close()  ###关闭上一个页面

                            driver.switch_to.window(windows[-1])

                            time.sleep(2.5)




                    else :

                        print(Count_Row)

                        Sheet.cell(row=Count_Row, column=11).value = '####'

                        Count_Row = Count_Row + 1

                        Sheet.cell(row=1, column=31).value = Count_Row

                        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                        Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                        driver.get(url)

                        time.sleep(2.5)

                        windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                        # driver.close()  ###关闭上一个页面

                        driver.switch_to.window(windows[-1])

                        time.sleep(2.5)




                else :

                    print(Count_Row)

                    Sheet.cell(row=Count_Row, column=11).value = '####'

                    Count_Row = Count_Row + 1

                    Sheet.cell(row=1, column=31).value = Count_Row

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                    Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                    driver.get(url)

                    time.sleep(2.5)

                    windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                    # driver.close()  ###关闭上一个页面

                    driver.switch_to.window(windows[-1])

                    time.sleep(2.5)




            else :

                print(Count_Row)

                Sheet.cell(row=Count_Row, column=11).value = '####'

                Count_Row = Count_Row + 1

                Sheet.cell(row=1, column=31).value = Count_Row

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)  ###搜索框中要输入的店铺

                driver.get(url)

                time.sleep(2.5)

                windows = driver.window_handles  ####搜索当前所有打开的页面的句柄（识别用）

                # driver.close()  ###关闭上一个页面

                driver.switch_to.window(windows[-1])

                time.sleep(2.5)




        ########时间除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=5).value

        while Count_Row < Max_Row + 1:

            if None_Count == '####':

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=5).value

                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:
                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)

            if None_Count != '####':

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=5).value

        time.sleep(0.3)

        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")
        print("时间除空结束。")

        ########时间除空插件

        ########地址除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=11).value

        while Count_Row < Max_Row + 1:

            if None_Count == '####':

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=11).value

                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:
                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)

            if None_Count != '####':

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=11).value

        time.sleep(0.3)

        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")
        print("地址除空结束。")

        ########地址除空插件

        ########电话除空插件

        Count_Row = 1

        None_Count = Sheet.cell(row=Count_Row, column=12).value

        while Count_Row < Max_Row + 1:

            if None_Count == '####':

                Sheet.delete_rows(Count_Row)

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Count_Row = Count_Row

                None_Count = Sheet.cell(row=Count_Row, column=12).value

                Total_Row = 1

                Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                while Total_Quantity != None:
                    Total_Row = Total_Row + 1

                    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

                Max_Row = Total_Row - 1

                print(Max_Row)

            if None_Count != '####':

                Count_Row = Count_Row + 1

                None_Count = Sheet.cell(row=Count_Row, column=12).value

        time.sleep(0.3)

        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")
        print("电话除空结束。")

        ########电话除空插件




    except :

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        driver.quit()

        #time.sleep(10)

        #goto.begin

        #restart_program()




Total()
###NT























































