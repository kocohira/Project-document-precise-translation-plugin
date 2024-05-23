#! /usr/bin/env python3
#! /usr/bin/env python
import goto
from dominate.tags import label
from goto import with_goto




@with_goto
def Total() :

    label.begin

    try :

        from selenium import webdriver
        import json
        import re       ###正则表达式用
        from openpyxl import load_workbook          ###Excel写入用
        from datetime import datetime
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        import nltk                                 ###自然语言处理用
        import time
        from selenium.webdriver.common.action_chains import ActionChains






        #实例化谷歌设置选项
        option = webdriver.ChromeOptions()
        #添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
        #option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
        option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
        #option.add_argument("--headless")


        experimentalFlags = [
            #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
            'same-site-by-default-cookies@2',
            #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
            'cookies-without-same-site-must-be-secure@2',]
        chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
        option.add_experimental_option('localState', chromeLocalStatePrefs)
        #初始化driver
        driver = webdriver.Chrome(options=option)








        Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

        Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表




        Count = 1

        Output_Row = Sheet.cell(row=1, column=32).value         ####出错从这里修改

        Output_Column_2_1_1 = 19

        Output_Column_Week = 26

        Polld_Source = Sheet.cell(row=Output_Row, column=1).value

        Single_OR_Double = Sheet.cell(row=Output_Row, column=Output_Column_2_1_1).value

        Week_Source = Sheet.cell(row=Output_Row, column=26).value

        Telephone_Source = Sheet.cell(row=Output_Row, column=12).value


        ########Excel计数插件A

        Total_Row_Sheet = 1

        Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        while Total_Quantity != None :

            Total_Row_Sheet = Total_Row_Sheet + 1

            Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        Max_Row_Sheet = Total_Row_Sheet - 1

        print(Max_Row_Sheet)

        ########Excel计数插件A











        url = 'https://eo.dianping.com/epassport/bookauthmanage'

        driver.get(url)

        cookies1_txt = '''[
        {
            "domain": ".dianping.com",
            "expirationDate": 1669735862,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_ga",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.2.723816111.1603966407",
            "id": 1
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606663922,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_gat_gtag_UA_111805100_1",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 2
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606750262,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_gid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "GA1.2.544243607.1606512602",
            "id": 3
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1635502415,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_hc.v",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "25c06052-ec24-b321-75e2-df2018908df6.1603966415",
            "id": 4
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1698574407,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
            "id": 5
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1698574407,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk_cuid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
            "id": 6
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1606665662,
            "hostOnly": false,
            "httpOnly": false,
            "name": "_lxsdk_s",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "176149fb99b-60e-a06-449%7C%7C1",
            "id": 7
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094172,
            "hostOnly": false,
            "httpOnly": false,
            "name": "aburl",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1",
            "id": 8
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1607268580.106885,
            "hostOnly": false,
            "httpOnly": false,
            "name": "bsid",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
            "id": 9
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1635761147.17086,
            "hostOnly": false,
            "httpOnly": false,
            "name": "ctu",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "9f07eccced25cf82543a98c2d6e0235399e464f43265dd96bf67e9790d7d49ed",
            "id": 10
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094249,
            "hostOnly": false,
            "httpOnly": false,
            "name": "cy",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "2374",
            "id": 11
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659.094289,
            "hostOnly": false,
            "httpOnly": false,
            "name": "cye",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "osaka",
            "id": 12
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1607268580.107027,
            "hostOnly": false,
            "httpOnly": false,
            "name": "edper",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "_oEUhHyIygGOLp4mrY-pfZigXt7jy32YqB47hMazKTNfp_uM5NzsH1A0FAodtX57SApkVDHB7EU3jPrA_Yt-Xg",
            "id": 13
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1637166659,
            "hostOnly": false,
            "httpOnly": false,
            "name": "Hm_lvt_602b80cf8079ae6591966cc70a3940e7",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "1604647683,1604835193,1605286882,1605628795",
            "id": 14
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1667994073.957559,
            "hostOnly": false,
            "httpOnly": false,
            "name": "s_ViewType",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "10",
            "id": 15
        },
        {
            "domain": ".dianping.com",
            "expirationDate": 1636371292.046143,
            "hostOnly": false,
            "httpOnly": false,
            "name": "ua",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "dpuser_83084044340",
            "id": 16
        },
        {
            "domain": ".eo.dianping.com",
            "expirationDate": 1607268662,
            "hostOnly": false,
            "httpOnly": false,
            "name": "os-agent",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": false,
            "storeId": "0",
            "value": "",
            "id": 17
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "JSESSIONID",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "1715B488C3EDE3BE27F67B2C5782285E",
            "id": 18
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "logan_custom_report",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "",
            "id": 19
        },
        {
            "domain": "eo.dianping.com",
            "hostOnly": true,
            "httpOnly": false,
            "name": "logan_session_token",
            "path": "/",
            "sameSite": "unspecified",
            "secure": false,
            "session": true,
            "storeId": "0",
            "value": "jr2ykj71td5ykdn3ngch",
            "id": 20
        }
        ]'''          ###cookies内容

        cookies1 = json.loads(cookies1_txt)###cookies json化

        driver.delete_all_cookies()      ###清除原有的cookies

        for cook in cookies1 :            ###cookies去sameSite开始
            try :
                cook.pop('sameSite')
            except :
                pass
            driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址

        #time.sleep(2)

        driver.get(url)                  ###重新加载网址







        while Output_Row < Max_Row_Sheet + 1 :               ###这里将来要改循环

            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[1]/div[1]/div/div[2]/div/span/input")))  ###等待输入餐厅的POLLD号码框加载完毕

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[1]/div[1]/div/div[2]/div/span/input").send_keys(Polld_Source)                 ###输入餐厅的POLLD号码

            time.sleep(0.8)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form/div[3]/div/button").click()                                                     ###点击查询

            time.sleep(1.2)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/div/div/div/div/div/div/table/tbody/tr/td[8]/div/span[1]/a").click()                 ###加载完毕后，点击授权按钮

            time.sleep(0.8)

            Handle_1 = driver.current_url

            time.sleep(1.2)




            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[1]/div[2]/div/span/div/div[1]/label/span[1]/input").click()              ###点击‘电话预定’选项

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[1]/div[2]/div/span/div/div[2]/input").send_keys('81')                    ###输入国际区号

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[1]/div[2]/div/span/div/div[4]/input").send_keys(Telephone_Source)              ###输入餐厅联系电话

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()                                      ###点击‘新增’按钮，增加一行营业时间

            time.sleep(0.7)








            if Single_OR_Double == None :

                Time_1_1_1 = int(Sheet.cell(row=Output_Row, column=15).value) + 1

                Time_1_1_1 = str(Time_1_1_1)

                Time_1_1_2_Source = Sheet.cell(row=Output_Row, column=16).value

                if Time_1_1_2_Source == '00':

                    Time_1_1_2 = str(1)

                else:

                    Time_1_1_2 = str(2)

                Time_1_2_1 = int(Sheet.cell(row=Output_Row, column=17).value) + 1

                Time_1_2_1 = str(Time_1_2_1)

                Time_1_2_2_Source = Sheet.cell(row=Output_Row, column=18).value

                if Time_1_2_2_Source == '00':

                    Time_1_2_2 = str(1)

                else:

                    Time_1_2_2 = str(2)




                if Week_Source == str(0) or Week_Source == 0 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1).click()

                    time.sleep(0.7)

                    Hover_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(1) or Week_Source == 1 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu  ant-select-dropdown-menu-root ant-select-dropdown-menu-vertical']/li[2]").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1).click()

                    time.sleep(0.7)

                    Hover_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(2) or Week_Source == 2 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)


                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu-item ant-select-dropdown-menu-item-active' and text()='周一']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周三']")[1].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(3) or Week_Source == 3 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周二']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周四']")[1].click()  ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(4) or Week_Source == 4 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周三']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周五']")[1].click()  ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(5) or Week_Source == 5 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周四']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周六']")[1].click()  ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(6) or Week_Source == 6 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周五']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周日']")[2].click()  ###这里不知道为什么是【2】

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(7) or Week_Source == 7 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu  ant-select-dropdown-menu-root ant-select-dropdown-menu-vertical']/li[6]").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1).click()

                    time.sleep(0.7)

                    Hover_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                if Week_Source == str(8) or Week_Source == 8 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu  ant-select-dropdown-menu-root ant-select-dropdown-menu-vertical']/li[5]").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1).click()

                    time.sleep(0.7)

                    Hover_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()








            if Single_OR_Double != None :

                Time_1_1_1 = int(Sheet.cell(row=Output_Row, column=15).value) + 1

                Time_1_1_1 = str(Time_1_1_1)

                Time_1_1_2_Source = Sheet.cell(row=Output_Row, column=16).value

                if Time_1_1_2_Source == '00':

                    Time_1_1_2 = str(1)

                else:

                    Time_1_1_2 = str(2)

                Time_1_2_1 = int(Sheet.cell(row=Output_Row, column=17).value) + 1

                Time_1_2_1 = str(Time_1_2_1)

                Time_1_2_2_Source = Sheet.cell(row=Output_Row, column=18).value

                if Time_1_2_2_Source == '00':

                    Time_1_2_2 = str(1)

                else:

                    Time_1_2_2 = str(2)

                Time_2_1_1 = int(Sheet.cell(row=Output_Row, column=19).value) + 1

                Time_2_1_1 = str(Time_2_1_1)

                Time_2_1_2_Source = Sheet.cell(row=Output_Row, column=20).value

                if Time_2_1_2_Source == '00':

                    Time_2_1_2 = str(1)

                else:

                    Time_2_1_2 = str(2)

                Time_2_2_1 = int(Sheet.cell(row=Output_Row, column=21).value) + 1

                Time_2_2_1 = str(Time_2_2_1)

                Time_2_2_2_Source = Sheet.cell(row=Output_Row, column=22).value

                if Time_2_2_2_Source == '00':

                    Time_2_2_2 = str(1)

                else:

                    Time_2_2_2 = str(2)




                if Week_Source == str(0) or Week_Source == 0 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)





                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1).click()

                    time.sleep(0.7)

                    Hover_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(1) or Week_Source == 1 :

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button")))

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[1]/div/div/div")))  ###等待输入餐厅的POLLD号码框加载完毕

                    XinZeng_2_1 = driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[1]/div/div/div").click()

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//*[text()='周二']")))  ###等待输入餐厅的POLLD号码框加载完毕

                    driver.find_element_by_xpath("//*[text()='周二']").click()  ###一会给出解释

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div")))  ###等待输入餐厅的POLLD号码框加载完毕

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[1]/div/div/div").click()

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH,"//*[text()='周二']")))  ###等待输入餐厅的POLLD号码框加载完毕

                    driver.find_elements_by_xpath("//*[text()='周二']")[2].click()  ###一会给出解释

                    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.XPATH, "//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input")))  ###等待输入餐厅的POLLD号码框加载完毕




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(2) or Week_Source == 2 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu-item ant-select-dropdown-menu-item-active' and text()='周一']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@class='ant-select-dropdown-menu-item ant-select-dropdown-menu-item-active' and text()='周一']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周三']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周三']")[4].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(3) or Week_Source == 3 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周二']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周二']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周四']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周四']")[4].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(4) or Week_Source == 4 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周三']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周三']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周五']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周五']")[4].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(5) or Week_Source == 5 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周四']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周四']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周六']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周六']")[4].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(6) or Week_Source == 6 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周五']").click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周五']")[2].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周日']")[4].click()            ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[1]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周日']")[6].click()            ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[3]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[4]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(7) or Week_Source == 7 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周六']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周六']")[2].click()  ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()




                if Week_Source == str(8) or Week_Source == 8 :

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/button").click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[text()='周五']").click()  ###一会给出解释

                    time.sleep(0.7)

                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[3]/div/div/div").click()

                    time.sleep(0.7)

                    driver.find_elements_by_xpath("//*[text()='周五']")[2].click()  ###一会给出解释

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_1_1)

                    driver.find_element_by_xpath(Timetextbox_1_1_1).click()

                    time.sleep(0.7)

                    Hover_1_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_1_2)

                    driver.find_element_by_xpath(Timetextbox_1_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[1]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_1_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_1_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_1_2_1)

                    driver.find_element_by_xpath(Timetextbox_1_2_1).click()

                    time.sleep(0.7)

                    Hover_1_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_1_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_1_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_1_2_2)

                    driver.find_element_by_xpath(Timetextbox_1_2_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[5]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_1_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_1_1)

                    driver.find_element_by_xpath(Timetextbox_2_1_1).click()

                    time.sleep(0.7)

                    Hover_2_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_1_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_1_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_1_2)

                    driver.find_element_by_xpath(Timetextbox_2_1_2).click()

                    time.sleep(0.7)




                    driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[5]/div[2]/div/span/div[2]/div/div/div/span/div[7]/span/input").click()

                    time.sleep(0.7)

                    Hover_2_2_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

                    ActionChains(driver).move_to_element(Hover_2_2_1).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_1 = "//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[{}]".format(Time_2_2_1)

                    driver.find_element_by_xpath(Timetextbox_2_2_1).click()

                    time.sleep(0.7)

                    Hover_2_2_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

                    ActionChains(driver).move_to_element(Hover_2_2_2).perform()

                    time.sleep(0.7)

                    Timetextbox_2_2_2 = "//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[{}]".format(Time_2_2_2)

                    driver.find_element_by_xpath(Timetextbox_2_2_2).click()








            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[6]/div[2]/div/span/div/div[1]/span/input").send_keys('1')                ###提前预定时间设为’1‘天

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[6]/div[2]/div/span/div/div[2]/span/input").click()                       ###点击‘提前预定时间点’框

            time.sleep(0.7)

            Hover_3_1_1 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]")

            ActionChains(driver).move_to_element(Hover_3_1_1).perform()

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[1]/ul[1]/li[21]").click()                                                    ###选中’20‘时

            time.sleep(0.7)

            Hover_3_1_2 = driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]")

            ActionChains(driver).move_to_element(Hover_3_1_2).perform()

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@class='ant-time-picker-panel-combobox']/div[2]/ul[1]/li[1]").click()                                                         ###选中’00‘分，

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[7]/div[2]/div/span/div/div[1]/input").send_keys('1')                     ###预定人数Min限制设为’1‘人

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[7]/div[2]/div/span/div/div[3]/input").send_keys('10')                    ###预定人数Max限制设为’10‘人

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@id='app']/div/div/div[2]/div[2]/div/div/form[2]/div[8]/button").click()                                                      ###点击’提交授权‘



            ##################下面直到最后完成前不准打开，因为会直接提交数据

            time.sleep(0.7)

            driver.find_element_by_xpath("//*[@class='ant-modal-footer']/div[1]/button[2]").click()

            #driver.find_elements_by_xpath("//*[@class='ant-btn ant-btn-primary']").click()

            ##################上面直到最后完成前不准打开，因为会直接提交数据




            time.sleep(4)

            Handle_2 = driver.current_url

            if Handle_1 == Handle_2 :

                Result = '否'

                Sheet.cell(row=Output_Row, column=28).value = Result

                print(Output_Row)

                Count = Count + 1

                Output_Row = Output_Row + 1

                Sheet.cell(row=1, column=32).value = Output_Row

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Polld_Source = Sheet.cell(row=Output_Row, column=1).value

                Single_OR_Double = Sheet.cell(row=Output_Row, column=Output_Column_2_1_1).value

                Week_Source = Sheet.cell(row=Output_Row, column=26).value

                Telephone_Source = Sheet.cell(row=Output_Row, column=12).value

            if Handle_1 != Handle_2 :

                Result = '是'

                Sheet.cell(row=Output_Row, column=28).value = Result

                print(Output_Row)

                Count = Count + 1

                Output_Row = Output_Row + 1

                Sheet.cell(row=1, column=32).value = Output_Row

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

                Polld_Source = Sheet.cell(row=Output_Row, column=1).value

                Single_OR_Double = Sheet.cell(row=Output_Row, column=Output_Column_2_1_1).value

                Week_Source = Sheet.cell(row=Output_Row, column=26).value

                Telephone_Source = Sheet.cell(row=Output_Row, column=12).value






            if Output_Row < Max_Row_Sheet + 1 :            ###这里将来要改循环

                time.sleep(0.7)

                driver.execute_script("window.open('{}');".format(url))

                time.sleep(0.7)

                driver.close()

                time.sleep(0.7)

                driver.switch_to.window(driver.window_handles[-1])

                time.sleep(0.7)




        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        driver.quit()

        print('录入完成')
        print('录入完成')
        print('录入完成')
        print('录入完成')
        print('录入完成')
        print('录入完成')
        print('录入完成')
        print('录入完成')




    except :

        # ! /usr/bin/env python3
        # ! /usr/bin/env python

        #import time,win32api,win32con
        #import os
        #import sys
        #import 3_A_食べログ餐厅信息搜寻系统

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        driver.quit()

        time.sleep(8)

        goto.begin

        #def restart_program() :

            #print('重启程序中...')

            #python = sys.executable

            #os.execl(python,python,*sys.argv)

        #os.popen('D:\应用方面\python project2\A_3_食べログ餐厅信息搜寻系统.py')
        #restart_program()




Total()
###NT


