from nltk.tokenize import word_tokenize
import re
from openpyxl import load_workbook          ###Excel写入用
import time






Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表

Excel_Dictionary = load_workbook(r'C:\Users\kocohira\Desktop\Special\辖区&日期_字典.xlsx') ###打开字典Excel

Sheet_Dictionary = Excel_Dictionary['字典']

Row_Time_TBLG = 236         ###出错从这里修改

Column_Time_TBLG = 23

Row_Time_Dictionary = 2

Column_Time_Dictionary = 9

Week_TBLG = str((Sheet.cell(row=Row_Time_TBLG, column=Column_Time_TBLG)).value)

Week_Dictionary = str((Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=Column_Time_Dictionary)).value)




########Excel计数插件A

Total_Row_Sheet = 1

Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

while Total_Quantity != None :

    Total_Row_Sheet = Total_Row_Sheet + 1

    Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

Max_Row_Sheet = Total_Row_Sheet - 1

print(Max_Row_Sheet)

########Excel计数插件A




########Excel计数插件B(字典)

Total_Row_Sheet_Dictionary = 1

Total_Quantity = Sheet_Dictionary.cell(row=Total_Row_Sheet_Dictionary, column=9).value

while Total_Quantity != None :

    Total_Row_Sheet_Dictionary = Total_Row_Sheet_Dictionary + 1

    Total_Quantity = Sheet_Dictionary.cell(row=Total_Row_Sheet_Dictionary, column=9).value

Max_Row_Sheet_Dictionary = Total_Row_Sheet_Dictionary - 1

print(Max_Row_Sheet_Dictionary)

########Excel计数插件B(字典)











###

while Row_Time_TBLG < Max_Row_Sheet + 1 :

    while re.match(Week_Dictionary,Week_TBLG) == None and Row_Time_Dictionary < Max_Row_Sheet_Dictionary :

        Row_Time_Dictionary = Row_Time_Dictionary + 1

        Week_Dictionary = Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=Column_Time_Dictionary).value

    if Row_Time_Dictionary >= Max_Row_Sheet_Dictionary :

        Sheet.cell(row=Row_Time_TBLG, column=26).value = '0'

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        print(Row_Time_TBLG)

        Row_Time_TBLG = Row_Time_TBLG + 1

        Row_Time_Dictionary = 2

        Week_Dictionary = str((Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=Column_Time_Dictionary)).value)

        Week_TBLG = str((Sheet.cell(row=Row_Time_TBLG, column=Column_Time_TBLG)).value)

    if re.match(Week_Dictionary,Week_TBLG) != None:

        Sheet.cell(row = Row_Time_TBLG, column = 26).value = Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=11).value

        Sheet.cell(row=Row_Time_TBLG, column=4).value = Sheet.cell(row=Row_Time_TBLG, column=4).value + '  ' + Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=10).value

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        print(Row_Time_TBLG)

        Row_Time_TBLG = Row_Time_TBLG + 1

        Row_Time_Dictionary = 2

        Week_Dictionary = str((Sheet_Dictionary.cell(row=Row_Time_Dictionary, column=Column_Time_Dictionary)).value)

        Week_TBLG = str((Sheet.cell(row=Row_Time_TBLG, column=Column_Time_TBLG)).value)


print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
print('比对结束')
###NT






