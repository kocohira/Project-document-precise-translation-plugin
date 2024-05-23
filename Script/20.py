from selenium import webdriver
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import nltk                                 ###自然语言处理用
import time,win32api,win32con
import os
import sys



def restart_program() :

    print('重启程序中...')

    python = sys.executable

    os.execl(python,python,*sys.argv)




#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
#option.add_argument("--headless")

experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)



#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

Sheet = Excel['Main']  ###打开‘大阪’工作表

#Count_Row = int((Sheet.cell(row=1, column=31)).value)                             ###每次出错就从这里修改

#Restanrant_Search = str((Sheet.cell(row=Count_Row, column=2)).value)                ###搜索框中要输入的店铺

#Location_Search = str((Sheet.cell(row=Count_Row, column=9)).value)                  ###搜索框中要输入的区域




########Excel计数插件

#Total_Row = 1

#Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

#while Total_Quantity != None :

    #Total_Row = Total_Row + 1

    #Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

#Max_Row = Total_Row - 1             ###包含了表格第一行

#print(Max_Row)

########Excel计数插件








url = 'https://manager.starday.shop/?#/thirdsupplyproducts'



driver.get(url)

cookies2_txt = """[{"captcha": "e4886c33385ce84ef255bc198ddf75dc"}]"""          ###cookies内容

cookies2 = json.loads(cookies2_txt)###cookies json化

driver.delete_all_cookies()      ###清除原有的cookies

for cook in cookies2 :            ###cookies去sameSite开始
    try :
        cook.pop('sameSite')
    except :
        pass
    driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址
driver.get(url)                  ###重新加载网址
###NT
