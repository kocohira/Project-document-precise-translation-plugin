from selenium import webdriver
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import nltk                                 ###自然语言处理用




#实例化谷歌设置选项
option = webdriver.ChromeOptions()
#添加保持登录的数据路径：安装目录一般在C:\Users\黄\AppData\Local\Google\Chrome\User Data
#option.add_argument(r'user-data-dir=C:\Users\kocohira\AppData\Local\Google\Chrome\User Data2')
option.add_argument('user-agent="Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"')
experimentalFlags = [
    #【机器翻译官方说明】站点必须指定SameSite=None才能启用第三方使用。0默认1开启2关闭
    'same-site-by-default-cookies@2',
    #【机器翻译官方说明】如果设置了没有SameSite限制的cookie而没有Secure属性，则将拒绝该cookie。0默认1开启2关闭
    'cookies-without-same-site-must-be-secure@2',]
chromeLocalStatePrefs = { 'browser.enabled_labs_experiments' : experimentalFlags}
option.add_experimental_option('localState', chromeLocalStatePrefs)



#初始化driver
driver = webdriver.Chrome(options=option)








Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')  ###读取Excel比较表格

Sheet = Excel['大阪']  ###打开‘大阪’工作表

Row3 = 2

#Column3 = 2

#Row4 = 2

#Column4 = 9

Restanrant_Search = str((Sheet.cell(row=Row3, column=2)).value)                ###搜索框中要输入的店铺

Location_Search = str((Sheet.cell(row=Row3, column=9)).value)                  ###搜索框中要输入的区域










url = 'https://tabelog.com/osaka/A2701/A270405/27011913/'

driver.get(url)

cookies2_txt = """[
{
    "domain": ".dianping.com",
    "expirationDate": 1668961275,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_ga",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.723816111.1603966407",
    "id": 1
},
{
    "domain": ".dianping.com",
    "expirationDate": 1605889335,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gat_gtag_UA_111805100_1",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 2
},
{
    "domain": ".dianping.com",
    "expirationDate": 1605975675,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_gid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "GA1.2.369363227.1605818515",
    "id": 3
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635502415,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_hc.v",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "25c06052-ec24-b321-75e2-df2018908df6.1603966415",
    "id": 4
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606235459,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lx_utm",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "utm_source%3Dgoogle%26utm_medium%3Dorganic",
    "id": 5
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 6
},
{
    "domain": ".dianping.com",
    "expirationDate": 1698574407,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_cuid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "17573d91220c8-005c20ba1064d6-303464-1fa400-17573d91220c8",
    "id": 7
},
{
    "domain": ".dianping.com",
    "expirationDate": 1605890141,
    "hostOnly": false,
    "httpOnly": false,
    "name": "_lxsdk_s",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "175e6599fa8-5fc-f4c-324%7C%7C3",
    "id": 8
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094172,
    "hostOnly": false,
    "httpOnly": false,
    "name": "aburl",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1",
    "id": 9
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606486817.735597,
    "hostOnly": false,
    "httpOnly": false,
    "name": "bsid",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "KIm9nv3SR27k0fzg4zPvth8BIQcTEZwQUs_66vzjsOk4aFXKveYCZ8G0Eqb2E-QlrLHI_s_guYaocm60ZQ3UTg",
    "id": 10
},
{
    "domain": ".dianping.com",
    "expirationDate": 1635761147.17086,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ctu",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "9f07eccced25cf82543a98c2d6e0235399e464f43265dd96bf67e9790d7d49ed",
    "id": 11
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094249,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cy",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "2374",
    "id": 12
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659.094289,
    "hostOnly": false,
    "httpOnly": false,
    "name": "cye",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "osaka",
    "id": 13
},
{
    "domain": ".dianping.com",
    "expirationDate": 1606486817.735668,
    "hostOnly": false,
    "httpOnly": false,
    "name": "edper",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "KIm9nv3SR27k0fzg4zPvth8BIQcTEZwQUs_66vzjsOk4aFXKveYCZ8G0Eqb2E-QlrLHI_s_guYaocm60ZQ3UTg",
    "id": 14
},
{
    "domain": ".dianping.com",
    "expirationDate": 1637166659,
    "hostOnly": false,
    "httpOnly": false,
    "name": "Hm_lvt_602b80cf8079ae6591966cc70a3940e7",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "1604647683,1604835193,1605286882,1605628795",
    "id": 15
},
{
    "domain": ".dianping.com",
    "expirationDate": 1667994073.957559,
    "hostOnly": false,
    "httpOnly": false,
    "name": "s_ViewType",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "10",
    "id": 16
},
{
    "domain": ".dianping.com",
    "expirationDate": 1636371292.046143,
    "hostOnly": false,
    "httpOnly": false,
    "name": "ua",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "dpuser_83084044340",
    "id": 17
},
{
    "domain": ".eo.dianping.com",
    "expirationDate": 1606494075,
    "hostOnly": false,
    "httpOnly": false,
    "name": "os-agent",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": false,
    "storeId": "0",
    "value": "",
    "id": 18
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "JSESSIONID",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "E72C3035785E5032D843FDF05CA009A7",
    "id": 19
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_custom_report",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "",
    "id": 20
},
{
    "domain": "eo.dianping.com",
    "hostOnly": true,
    "httpOnly": false,
    "name": "logan_session_token",
    "path": "/",
    "sameSite": "unspecified",
    "secure": false,
    "session": true,
    "storeId": "0",
    "value": "fx6olnip3eu2qra64wfa",
    "id": 21
}
]"""          ###cookies内容

cookies2 = json.loads(cookies2_txt)###cookies json化

driver.delete_all_cookies()      ###清除原有的cookies

for cook in cookies2 :            ###cookies去sameSite开始
    try :
        cook.pop('sameSite')
    except :
        pass
    driver.add_cookie(cook)      ###cookies去sameSite结束，并将cookie添加入目标网址
driver.get(url)                  ###重新加载网址








WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='sk']")))          ###等待出现搜索店铺的搜索框加载完毕




driver.find_element_by_xpath("//*[@id='sk']").send_keys(Restanrant_Search)                          ###找到输入店铺的搜索框，输入店铺

driver.find_element_by_xpath("//*[@id='sa']").send_keys(Location_Search)                            ###找到输入区域的搜索框，输入区域

driver.find_element_by_xpath("//*[@id='js-global-search-btn']").click()                             ###找到搜索按钮并点击，窗口跳到下一个画面，但控制的页面还是上一个

WebDriverWait(driver,30).until(EC.presence_of_element_located((By.XPATH,"//*[@id='container']/div[15]/div[4]/div/div[6]/div[1]/div[2]/div/div[1]/div/div[1]/h4/a")))            ###等待第一个搜索结果加载完毕

driver.find_element_by_xpath("//*[@id='container']/div[15]/div[4]/div/div[6]/div[1]/div[2]/div/div[1]/div/div[1]/h4/a").click()         ###缺一个判断有无店铺结果的语句

windows = driver.window_handles                                                                     ####搜索当前所有打开的页面的句柄（识别用）

driver.close()                                                                                      ###关闭上一个页面

driver.switch_to.window(windows[-1])                                                                ###切换到要控制的新开的网页的句柄。





###在这里判断店铺有无电话，地址，营业时间


TBRG_Telephone = driver.find_element_by_xpath("//*[@class='rstinfo-table__tel-num']")      ###找到餐厅电话

Telephone_Sourse1 = TBRG_Telephone.text                                                      ###将餐厅电话文本化

#print(Telephone_Sourse1)

Telephone_Sourse2 = re.sub("\D","",Telephone_Sourse1)                   ###将餐厅电话只保留数字

print(Telephone_Sourse2)

#Telephone_Sourse3 = re.sub("[\u0800-\u4e00]+", "", Telephone_Sourse2)                             ###将餐厅电话去除日文平片假名

#Telephone_Process = Telephone_Sourse3.replace(' ','')                                         ###将餐厅电话去空格

Telephone_Judgement = Telephone_Sourse2.isdigit()                       ###判断处理过后的电话是不是只剩下数字

#print(Telephone_Judgement)

'''if Telephone_Judgement == True :

    TBRG_Location = driver.find_element_by_xpath("//*[@class='rstinfo-table__address']")

    Location_Sourse1 = TBRG_Location.text

    print(Location_Sourse1)

    if len(Location_Sourse1) != 0 :

        Sheet.cell(row=Row3, column=12).value = Telephone_Sourse2

        Sheet.cell(row=Row3, column=11).value = Location_Sourse1




        TRBG_Time = driver.find_element_by_xpath("//*[@id='rst-data-head']/table[1]/tbody/tr[7]/td/p[2]")

        Time_Source1 = TRBG_Time.text

        print(Time_Source1)'''

TBRG_Location = driver.find_element_by_xpath("//*[@class='rstinfo-table__address']")

TBRG_Time = driver.find_element_by_xpath("//*[@id='rst-data-head']/table[1]/tbody/tr[last()]/td/p[2]")






















    #    [:]
    #    [：]
    #    [：]
    #    [:]

#Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

#Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表

#Sheet.cell(row = 2,column = 8).value = text_Output                                      ###向规定单元格写入数据

#Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')











###//*[@id="rst-data-head"]/table[1]/tbody/tr[7]/td/p[4]
###//*[@id="rst-data-head"]/table[1]/tbody/tr[7]/td/p[4]###定休日的Xpath
###NT
