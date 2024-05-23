from nltk.tokenize import word_tokenize
import re
from openpyxl import load_workbook          ###Excel写入用
import time



Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']                                                                    ###打开‘大阪’工作表








########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件




Count_Row = 2               ###出错在这里修改

text = Sheet.cell(row=Count_Row, column=5).value








while Count_Row < Max_Row + 1 :

    Text_Pre1 = text.find('24hours')

    Text_Pre2 = text.find('二十四時間営業')

    Text_Pre3 = text.find('24 hours')

    Text_Pre4 = text.find('24時間営業')

    if Text_Pre1 != -1 or Text_Pre2 != -1 or Text_Pre3 != -1 or Text_Pre4 != -1 :

        print(Count_Row)

        Sheet.cell(row=Count_Row, column=15).value = '00'

        Sheet.cell(row=Count_Row, column=16).value = '00'

        Sheet.cell(row=Count_Row, column=17).value = 23

        Sheet.cell(row=Count_Row, column=18).value = 30

        Sheet.cell(row=Count_Row, column=4).value = '00:00-23:30'

        Count_Row = Count_Row + 1

        text = Sheet.cell(row=Count_Row, column=5).value

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




    else :

        text = text.replace('翌','')
        text = text.replace('１','1')
        text = text.replace('２','2')
        text = text.replace('３','3')
        text = text.replace('４','4')
        text = text.replace('５','5')
        text = text.replace('６','6')
        text = text.replace('７','7')
        text = text.replace('８','8')
        text = text.replace('９','9')
        text = text.replace('０','0')

        text = text.replace('L',' L')

        text = text.replace('.', '')

        text = text.replace('：',':')

        text = text.replace('時',':')

        text = text.replace('分','')

        text = text.replace('/',' ')

        text = text.replace('／',' ')

        text = text.replace('～','-')

        text = text.replace(' ～ ','-')

        text = text.replace('~','-')

        text = text.replace(' ~ ','-')

        text = text.replace('〜','-')

        text = text.replace(' 〜 ', '-')

        text = text.replace(' - ','-')

        text = text.replace(' ー ','-')

        text = text.replace('ー','-')

        text = re.sub("[\!\%\,\〒\,\，\。\(\)\（\）\［\］]", " ", text)

        #text = re.sub("[\u0065-\u0090]+", "", text)

        #text = re.sub("[\u0097-\u0122]+", "", text)

        text = re.sub("[\u0800-\u9fa5]+", " ", text)

        print(text)

        List_Source = word_tokenize(text,"english")

        print(List_Source)
        ###如果List_Source的列表为Null，直接删去店铺。或者在搜集信息的一开始就全部检查一遍也可以。

        New_List_Source = []

        New_List_Source_0 = []

        New_List_Source_00 = []

        for data in List_Source :

            if re.search('-',data) != None :

                New_List_Source_0.append(data)

        for data in New_List_Source_0 :

            JudgeBox = re.compile('[0-9]+')

            Judge = JudgeBox.findall(data)

            if len(Judge) != 0 :

                New_List_Source_00.append(data)

        for data in New_List_Source_00 :

            if data.count(':') >= 2 :

                New_List_Source.append(data)


        print(Count_Row)

        print('时间结果 ',New_List_Source)

        print('len长 ',len(New_List_Source))


        if len(New_List_Source) == 0:

            Sheet.cell(row=Count_Row, column=15).value = '####'

            Count_Row = Count_Row + 1

            text = Sheet.cell(row=Count_Row, column=5).value

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


        if len(New_List_Source) == 1 :

            Time_1 = New_List_Source[0]

            if Time_1.count('pm') == 0 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

            if Time_1.count('pm') == 1 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_1[2] = str(int(Time_1[2]) + int(12))

            if Time_1.count('pm') == 2:

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_1[0] = str(int(Time_1[0]) + int(12))

                Time_1[2] = str(int(Time_1[2]) + int(12))

            if Time_1[1] == '15' :

                Time_1[1] = '00'

            if Time_1[1] == '45' :

                Time_1[1] = '30'

            if Time_1[3] == '15' :

                Time_1[3] = '00'

            if Time_1[3] == '45' :

                Time_1[3] = '30'

            if int(Time_1[2]) - int(Time_1[0]) < 0 and 24 - int(Time_1[2]) >= 0:

                Time_1[2] = 23

                Time_1[3] = 30

            if int(Time_1[2]) >= 24 :

                Time_1[2] = 23

            if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                Sheet.cell(row=Count_Row, column=17).value = 23

                Sheet.cell(row=Count_Row, column=18).value = 30

                Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                Count_Row = Count_Row + 1

                text = Sheet.cell(row=Count_Row, column=5).value

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




            else :

                Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                Count_Row = Count_Row + 1

                text = Sheet.cell(row=Count_Row, column=5).value

                Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




        if len(New_List_Source) > 1 :

            Time_1 = New_List_Source[0]

            Time_2 = New_List_Source[1]

            if Time_1.count('pm') == 0 and Time_2.count('pm') == 0:

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

            if Time_1.count('pm') == 1 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

                Time_1[2] = str(int(Time_1[2]) + int(12))

                Time_2[0] = str(int(Time_2[0]) + int(12))

                Time_2[2] = str(int(Time_2[2]) + int(12))

            if Time_1.count('pm') == 0 and Time_2.count('pm') > 0 :

                Time_1 = re.sub("[A-Za-z\:\-]+", " ", Time_1)

                Time_1 = word_tokenize(Time_1, "english")

                Time_2 = re.sub("[A-Za-z\:\-]+", " ", Time_2)

                Time_2 = word_tokenize(Time_2, "english")

                Time_2[0] = str(int(Time_2[0]) + int(12))

                Time_2[2] = str(int(Time_2[2]) + int(12))




            if Time_1[1] == '15' :

                Time_1[1] = '00'

            if Time_1[1] == '45' :

                Time_1[1] = '30'

            if Time_1[3] == '15' :

                Time_1[3] = '00'

            if Time_1[3] == '45' :

                Time_1[3] = '30'

            if Time_2[1] == '15' :

                Time_2[1] = '00'

            if Time_2[1] == '45' :

                Time_2[1] = '30'

            if Time_2[3] == '15' :

                Time_2[3] = '00'

            if Time_2[3] == '45' :

                Time_2[3] = '30'

            if int(Time_1[2]) - int(Time_1[0]) <= 0 :

                Time_1[2] = '23'




            if int(Time_2[0]) - int(Time_1[2]) >= 0 :

                if int(Time_2[2]) - int(Time_2[0]) < 0 and 24 - int(Time_2[2]) >= 0:

                    Time_2[2] = 23

                    Time_2[3] = 30

                if int(Time_2[2]) >= 24 :

                    Time_2[2] = 23

                if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=19).value = Time_2[0]

                    Sheet.cell(row=Count_Row, column=20).value = Time_2[1]

                    Sheet.cell(row=Count_Row, column=21).value = 23

                    Sheet.cell(row=Count_Row, column=22).value = 30

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0]) + '  ' + str(New_List_Source[1])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




                else:

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=19).value = Time_2[0]

                    Sheet.cell(row=Count_Row, column=20).value = Time_2[1]

                    Sheet.cell(row=Count_Row, column=21).value = str(int(Time_2[2]) - 1)

                    Sheet.cell(row=Count_Row, column=22).value = Time_2[3]

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0]) + '  ' + str(New_List_Source[1])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            if int(Time_2[0]) - int(Time_1[2]) < 0 :

                if int(Time_1[2]) - int(Time_1[0]) < 0 and 24 - int(Time_1[2]) >= 0 :

                    Time_1[2] = 23

                    Time_1[3] = 30

                if int(Time_1[2]) >= 24 :

                    Time_1[2] = 23

                if Time_1[2] == 23 or int(Time_1[2]) == 23 :

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = 23

                    Sheet.cell(row=Count_Row, column=18).value = 30

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')




                else:

                    Sheet.cell(row=Count_Row, column=15).value = Time_1[0]

                    Sheet.cell(row=Count_Row, column=16).value = Time_1[1]

                    Sheet.cell(row=Count_Row, column=17).value = str(int(Time_1[2]) - 1)

                    Sheet.cell(row=Count_Row, column=18).value = Time_1[3]

                    Sheet.cell(row=Count_Row, column=4).value = str(New_List_Source[0])

                    Count_Row = Count_Row + 1

                    text = Sheet.cell(row=Count_Row, column=5).value

                    Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')


########除空插件

Count_Row = 1

None_Count = Sheet.cell(row=Count_Row, column=15).value

while Count_Row < Max_Row + 1 :

    if None_Count == '####' :

        Sheet.delete_rows(Count_Row)

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        Count_Row = Count_Row

        None_Count = Sheet.cell(row=Count_Row, column=15).value




        Total_Row_Sheet = 1

        Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        while Total_Quantity != None:

            Total_Row_Sheet = Total_Row_Sheet + 1

            Total_Quantity = Sheet.cell(row=Total_Row_Sheet, column=1).value

        Max_Row_Sheet = Total_Row_Sheet - 1

        print(Max_Row_Sheet)




    if None_Count != '####' :

        Count_Row = Count_Row + 1

        None_Count = Sheet.cell(row=Count_Row, column=15).value

time.sleep(0.3)

print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")
print("除空结束。")


########除空插件###NT










