from selenium import webdriver
import time
import json
import re       ###正则表达式用
from openpyxl import load_workbook          ###Excel写入用




Excel = load_workbook(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')           ###读取Excel比较表格

Sheet = Excel['大阪']




########Excel计数插件

Total_Row = 1

Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

while Total_Quantity != None :

    Total_Row = Total_Row + 1

    Total_Quantity = Sheet.cell(row=Total_Row, column=1).value

Max_Row = Total_Row - 1             ###包含了表格第一行

print(Max_Row)

########Excel计数插件




Input_Row = 2

text = Sheet.cell(row=Input_Row, column=14).value


while Input_Row < Max_Row + 1 :

    if text != None :

        if text.find('google') != -1 or text.find('profile') != -1 or text.find('login') != -1 :

            Sheet.cell(row=Input_Row, column=14).value = ''

            print(Input_Row)

            Input_Row = Input_Row + 1

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            text = Sheet.cell(row=Input_Row, column=14).value




        else :

            print(Input_Row)

            Input_Row = Input_Row + 1

            Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

            text = Sheet.cell(row=Input_Row, column=14).value




    else:

        print(Input_Row)

        Input_Row = Input_Row + 1

        Excel.save(r'C:\Users\kocohira\Desktop\Special\ExcelProcess.xlsx')

        text = Sheet.cell(row=Input_Row, column=14).value




print('Over.Progress')
###NT




